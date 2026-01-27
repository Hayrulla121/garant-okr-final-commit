import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import json
import uuid
import os
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

TRANSLATIONS = {
    "en": {
        "title": "OKR Performance Tracker",
        "performance_scale": "Performance Scale",
        "department": "Department",
        "department_name": "Department Name",
        "create_department": "➕ Create New Department",
        "select_department": "Select Department",
        "delete_department": "🗑️ Delete Department",
        "no_departments": "No departments. Create one first!",
        "create_objective": "➕ Create New Objective",
        "objective_name": "Objective Name",
        "objective": "Objective",
        "add_key_results": "Add Key Results",
        "kr_name": "KR Name",
        "kr_description": "Description (hover tooltip)",
        "kr_description_placeholder": "Enter meaning/description of this KR...",
        "type": "Type",
        "higher_better": "↑ Higher is better",
        "lower_better": "↓ Lower is better",
        "qualitative": " Qualitative (A/B/C/D/E)",
        "unit": "Unit",
        "thresholds": "Thresholds",
        "add_kr": "➕ Add KR",
        "added_krs": "Added Key Results",
        "remove": "Remove",
        "create": "✅ Create Objective",
        "enter_name_error": "Enter objective name and add at least one KR",
        "score": "🎯 Score",
        "add_kr_to_obj": "➕ Add KR to this Objective",
        "add": "➕ Add",
        "delete_objective": "🗑️ Delete Objective",
        "export_excel": "Export Excel",
        "import_excel": "Import Excel",
        "import_success": "Data imported successfully!",
        "import_error": "Import failed: Invalid file format",
        "import_warning": "Some data could not be imported",
        "save_data": "Save Data",
        "load_data": "📂 Load Data",
        "data_saved": "✅ Data saved!",
        "data_loaded": "✅ Data loaded!",
        "no_data": "No saved data found",
        "load_demo": "📋 Load Demo",
        "create_first": "👆 Create your first objective!",
        "language": "Language",
        "fact": "Fact",
        "actual": "Actual",
        "result": "Result",
        "results_breakdown": "Results Breakdown",
        "delete": "Delete",
        "edit_manage": "Edit & Manage",
        "no_objectives_yet": "No objectives in this department yet.",
        "enter_dept_name": "Please enter a department name",
        "key_result": "Key Result",
        "no_krs": "No Key Results. Add some below.",
        "delete_krs": "🗑️ Delete Key Results",
        "performance_level": "Performance Level",
        "below": "Below", "meets": "Meets", "good": "Good", "exceptional": "Exceptional",
        "view_grid": "Grid",
        "view_full": "Full",
        "all_departments": "All Departments",
        "overview": "Overview",
        "total_objectives": "Total Objectives",
        "average_score": "Average Score",
        "departments": "Departments",
        "view_mode": "View Mode",
        "actions": "Actions",
        "avg": "Avg",
        "toggle_sidebar": "Toggle sidebar",
        "value": "Value",
        "weight": "Weight",
        "objective_weight": "Objective Weight (%)",
        "kr_weight": "KR Weight (%)",
        "weights_warning": "⚠️ Weights should sum to 100%",
        "weights_total": "Total",
        "qualitative_grade": "Grade",
        "grade_a": "A - Exceptional",
        "grade_b": "B - Very Good",
        "grade_c": "C - Good",
        "grade_d": "D - Meets",
        "grade_e": "E - Below",
        "weighted_score": "Weighted Score",
        "dept_weighted_avg": "Dept. Weighted Average",
        "weighted_formula": "Weighted Formula",
        "kr_contribution": "KR Contribution",
        "obj_contribution": "Objective Contribution",
        "formula_breakdown": "Formula Breakdown",
        "edit_kr": "Edit KR",
        "edit_krs": "✏️ Edit Key Results",
        "edit_objective": "✏️ Edit Objective",
        "update": "Update",
        "score_level_settings": "Score Level Settings",
        "configure_score_levels": "Configure Score Levels",
        "score_range": "Score Range",
        "min_score": "Min Score",
        "max_score": "Max Score",
        "performance_levels": "Performance Levels",
        "add_level": "Add Level",
        "delete_level": "Delete Level",
        "level_name": "Level Name",
        "level_threshold": "Threshold",
        "level_color": "Color",
        "grade_mapping": "Qualitative Grade Mapping",
        "save_settings": "Save Settings",
        "settings_saved": "Settings saved!",
        "invalid_config": "Invalid configuration",
        "reset_defaults": "Reset to Defaults",
        "cancel": "Cancel",
    },
    "ru": {
        "title": "OKR Трекер",
        "performance_scale": "Шкала Оценки",
        "department": "Департамент",
        "department_name": "Название Департамента",
        "create_department": "➕ Создать Департамент",
        "select_department": "Выберите Департамент",
        "delete_department": "🗑️ Удалить Департамент",
        "no_departments": "Нет департаментов. Создайте сначала!",
        "create_objective": "➕ Создать Цель",
        "objective_name": "Название Цели",
        "objective": "Цель",
        "add_key_results": "Добавить Ключевые Результаты",
        "kr_name": "Название KR",
        "kr_description": "Описание (подсказка)",
        "kr_description_placeholder": "Введите смысл/описание KR...",
        "type": "Тип",
        "higher_better": "↑ Больше лучше",
        "lower_better": "↓ Меньше лучше",
        "qualitative": " Качественный (A/B/C/D/E)",
        "unit": "Единица",
        "thresholds": "Пороги",
        "add_kr": "➕ Добавить KR",
        "added_krs": "Добавленные KR",
        "remove": "Удалить",
        "create": "✅ Создать",
        "enter_name_error": "Введите название и добавьте KR",
        "score": "🎯 Оценка",
        "add_kr_to_obj": "➕ Добавить KR",
        "add": "➕ Добавить",
        "delete_objective": "🗑️ Удалить Цель",
        "export_excel": "Экспорт Excel",
        "import_excel": "Импорт Excel",
        "import_success": "Данные успешно импортированы!",
        "import_error": "Ошибка импорта: неверный формат файла",
        "import_warning": "Некоторые данные не удалось импортировать",
        "save_data": "Сохранить",
        "load_data": "📂 Загрузить",
        "data_saved": "✅ Сохранено!",
        "data_loaded": "✅ Загружено!",
        "no_data": "Нет данных",
        "load_demo": "📋 Демо",
        "create_first": "👆 Создайте цель!",
        "language": "Язык",
        "fact": "Факт",
        "actual": "Фактический",
        "result": "Результат",
        "results_breakdown": "Разбивка результатов",
        "delete": "Удалить",
        "edit_manage": "Редактировать",
        "no_objectives_yet": "В этом департаменте пока нет целей.",
        "enter_dept_name": "Введите название департамента",
        "key_result": "Ключевой Результат",
        "no_krs": "Нет KR.",
        "delete_krs": "🗑️ Удалить Ключевые Результаты",
        "performance_level": "Уровень Производительности",
        "below": "Ниже ожидаемого", "meets": "На уровне ожиданий", "good": "Хорошо",
        "exceptional": "Исключительно",
        "view_grid": "Сетка",
        "view_full": "Полный",
        "all_departments": "Все Департаменты",
        "overview": "Обзор",
        "total_objectives": "Всего целей",
        "average_score": "Средняя оценка",
        "departments": "Департаменты",
        "view_mode": "Режим просмотра",
        "actions": "Действия",
        "avg": "Сред",
        "toggle_sidebar": "Скрыть/показать панель",
        "value": "Значение",
        "weight": "Вес",
        "objective_weight": "Вес цели (%)",
        "kr_weight": "Вес KR (%)",
        "weights_warning": "⚠️ Веса должны составлять 100%",
        "weights_total": "Итого",
        "qualitative_grade": "Оценка",
        "grade_a": "A - Исключительно",
        "grade_b": "B - Очень хорошо",
        "grade_c": "C - Хорошо",
        "grade_d": "D - Соответствует",
        "grade_e": "E - Ниже",
        "weighted_score": "Взвешенная оценка",
        "dept_weighted_avg": "Взвеш. среднее отдела",
        "weighted_formula": "Взвешенная формула",
        "kr_contribution": "Вклад KR",
        "obj_contribution": "Вклад цели",
        "formula_breakdown": "Разбивка формулы",
        "edit_kr": "Редактировать KR",
        "edit_krs": "✏️ Редактировать Ключевые Результаты",
        "edit_objective": "✏️ Редактировать Цель",
        "update": "Обновить",
        "score_level_settings": "Настройки уровней оценки",
        "configure_score_levels": "Настроить уровни оценки",
        "score_range": "Диапазон оценок",
        "min_score": "Мин. оценка",
        "max_score": "Макс. оценка",
        "performance_levels": "Уровни производительности",
        "add_level": "Добавить уровень",
        "delete_level": "Удалить уровень",
        "level_name": "Название уровня",
        "level_threshold": "Порог",
        "level_color": "Цвет",
        "grade_mapping": "Сопоставление оценок",
        "save_settings": "Сохранить настройки",
        "settings_saved": "Настройки сохранены!",
        "invalid_config": "Неверная конфигурация",
        "reset_defaults": "Сбросить по умолчанию",
        "cancel": "Отмена",
    },
    "uz": {
        "title": "OKR Трекер",
        "performance_scale": "Баҳолаш Шкаласи",
        "department": "Департамент",
        "department_name": "Департамент Номи",
        "create_department": "➕ Департамент Яратиш",
        "select_department": "Департамент Танланг",
        "delete_department": "🗑️ Департамент Ўчириш",
        "no_departments": "Департамент йўқ. Аввал яратинг!",
        "create_objective": "➕ Мақсад Яратиш",
        "objective_name": "Мақсад Номи",
        "objective": "Мақсад",
        "add_key_results": "Калит Натижалар Қўшиш",
        "kr_name": "KR Номи",
        "kr_description": "Тавсиф (кўрсатма)",
        "kr_description_placeholder": "KR маъносини киритинг...",
        "type": "Тури",
        "higher_better": "↑",
        "lower_better": "↓",
        "qualitative": " Сифат (A/B/C/D/E)",
        "unit": "Бирлик",
        "thresholds": "Чегаралар",
        "add_kr": "➕ KR Қўшиш",
        "added_krs": "Қўшилган KR",
        "remove": "Ўчириш",
        "create": "✅ Яратиш",
        "enter_name_error": "Ном ва KR киритинг",
        "score": "🎯 Баҳо",
        "add_kr_to_obj": "➕ KR Қўшиш",
        "add": "➕ Қўшиш",
        "delete_objective": "🗑️ Ўчириш",
        "export_excel": "Excel Экспорт",
        "import_excel": "Excel Импорт",
        "import_success": "Маълумотлар муваффақиятли импорт қилинди!",
        "import_error": "Импорт хатоси: нотўғри файл формати",
        "import_warning": "Баъзи маълумотларни импорт қилиб бўлмади",
        "save_data": "Сақлаш",
        "load_data": "📂 Юклаш",
        "data_saved": "✅ Сақланди!",
        "data_loaded": "✅ Юкланди!",
        "no_data": "Маълумот йўқ",
        "load_demo": "📋 Демо",
        "create_first": "👆 Мақсад яратинг!",
        "language": "Тил",
        "fact": "Факт",
        "actual": "Ҳақиқий",
        "result": "Натижа",
        "results_breakdown": "Натижалар тафсилоти",
        "delete": "Ўчириш",
        "edit_manage": "Таҳрирлаш",
        "no_objectives_yet": "Бу департаментда ҳали мақсадлар йўқ.",
        "enter_dept_name": "Департамент номини киритинг",
        "key_result": "Калит Натижа",
        "no_krs": "KR йўқ.",
        "delete_krs": "🗑️ Калит Натижаларни Ўчириш",
        "performance_level": "Самарадорлик Даражаси",
        "below": "Ёмон", "meets": "Кутилган", "good": "Яхши", "exceptional": "Фантастик",
        "view_grid": "Тўр",
        "view_full": "Тўлиқ",
        "all_departments": "Барча Департаментлар",
        "overview": "Умумий кўриниш",
        "total_objectives": "Жами мақсадлар",
        "average_score": "Ўртача баҳо",
        "departments": "Департаментлар",
        "view_mode": "Кўриш режими",
        "actions": "Амаллар",
        "avg": "Ўрт",
        "toggle_sidebar": "Панелни яшириш/кўрсатиш",
        "value": "Қиймат",
        "weight": "Вазн",
        "objective_weight": "Мақсад вазни (%)",
        "kr_weight": "KR вазни (%)",
        "weights_warning": "⚠️ Вазнлар 100% бўлиши керак",
        "weights_total": "Жами",
        "qualitative_grade": "Баҳо",
        "grade_a": "A - Фантастик",
        "grade_b": "B - Жуда яхши",
        "grade_c": "C - Яхши",
        "grade_d": "D - Кутилган",
        "grade_e": "E - Ёмон",
        "weighted_score": "Вазнли баҳо",
        "dept_weighted_avg": "Бўлим вазнли ўртача",
        "weighted_formula": "Вазнли формула",
        "kr_contribution": "KR ҳиссаси",
        "obj_contribution": "Мақсад ҳиссаси",
        "formula_breakdown": "Формула тафсилоти",
        "edit_kr": "KR таҳрирлаш",
        "edit_krs": "✏️ Калит Натижаларни Таҳрирлаш",
        "edit_objective": "✏️ Мақсадни Таҳрирлаш",
        "update": "Янгилаш",
        "score_level_settings": "Баҳо даражаси созламалари",
        "configure_score_levels": "Баҳо даражаларини созлаш",
        "score_range": "Баҳо оралиғи",
        "min_score": "Мин. баҳо",
        "max_score": "Макс. баҳо",
        "performance_levels": "Самарадорлик даражалари",
        "add_level": "Дарага қўшиш",
        "delete_level": "Даражани ўчириш",
        "level_name": "Дарага номи",
        "level_threshold": "Чегара",
        "level_color": "Ранг",
        "grade_mapping": "Баҳо хариталаш",
        "save_settings": "Созламаларни сақлаш",
        "settings_saved": "Созламалар сақланди!",
        "invalid_config": "Нотўғри конфигурация",
        "reset_defaults": "Стандартга қайтариш",
        "cancel": "Бекор қилиш",
    }
}

LEVELS = {
    "below": {"min": 4.25, "max": 4.49, "color": "#d9534f"},
    "meets": {"min": 4.50, "max": 4.74, "color": "#f0ad4e"},
    "good": {"min": 4.75, "max": 4.99, "color": "#5cb85c"},
    "exceptional": {"min": 5.00, "max": 5.00, "color": "#1e7b34"},
}

# Qualitative grades mapping (A/B/C/D/E to scores) - NEW SCALE (max 5.00)
# NOTE: These are kept for backward compatibility but dynamic config is preferred
QUALITATIVE_GRADES = {
    "A": {"score": 5.00, "level": "exceptional"},
    "B": {"score": 4.75, "level": "good"},
    "C": {"score": 4.50, "level": "meets"},
    "D": {"score": 4.25, "level": "below"},
    "E": {"score": 4.25, "level": "below"},
}

# Default configuration for dynamic score levels
DEFAULT_SCORE_LEVELS_CONFIG = {
    "min_score": 4.25,
    "max_score": 5.00,
    "levels": [
        {
            "key": "below",
            "order": 0,
            "threshold": 4.25,
            "color": "#d9534f",
            "names": {"en": "Below", "ru": "Ниже ожидаемого", "uz": "Ёмон"}
        },
        {
            "key": "meets",
            "order": 1,
            "threshold": 4.50,
            "color": "#f0ad4e",
            "names": {"en": "Meets", "ru": "На уровне ожиданий", "uz": "Кутилган"}
        },
        {
            "key": "good",
            "order": 2,
            "threshold": 4.75,
            "color": "#5cb85c",
            "names": {"en": "Good", "ru": "Хорошо", "uz": "Яхши"}
        },
        {
            "key": "exceptional",
            "order": 3,
            "threshold": 5.00,
            "color": "#1e7b34",
            "names": {"en": "Exceptional", "ru": "Исключительно", "uz": "Фантастик"}
        }
    ],
    "qualitative_mapping": {
        "A": "exceptional",
        "B": "good",
        "C": "meets",
        "D": "below",
        "E": "below"
    }
}

THEME = {
    "sidebar_bg": "#f5f7fa",
    "sidebar_border": "#e1e5eb",
    "main_bg": "#ffffff",
    "card_bg": "#ffffff",
    "card_border": "#e4e7ec",
    "card_shadow": "0 4px 12px rgba(0,0,0,0.08)",
    "text_primary": "#1a202c",
    "text_secondary": "#64748b",
    "accent": "#0066cc",
    "accent_light": "#e6f0ff",
    "header_bg": "linear-gradient(135deg, #1a365d 0%, #2c5282 100%)",
    "success": "#059669",
    "warning": "#d97706",
    "danger": "#dc2626",
}

DATA_FILE = "okr_data.json"


def t(key: str) -> str:
    lang = st.session_state.get('language', 'en')
    return TRANSLATIONS.get(lang, TRANSLATIONS['en']).get(key, key)


# ===== DYNAMIC SCORE LEVEL CONFIGURATION HELPERS =====

def get_score_levels_config() -> dict:
    """Get score levels configuration from session state or return default."""
    return st.session_state.get('score_levels_config', DEFAULT_SCORE_LEVELS_CONFIG)


def get_levels_dict() -> dict:
    """Build a LEVELS-compatible dict from dynamic config for backward compatibility."""
    config = get_score_levels_config()
    levels = sorted(config['levels'], key=lambda x: x['order'])
    result = {}

    for i, level in enumerate(levels):
        # Calculate max as threshold of next level minus 0.01, or same as threshold for last level
        if i < len(levels) - 1:
            max_val = round(levels[i + 1]['threshold'] - 0.01, 2)
        else:
            max_val = level['threshold']

        result[level['key']] = {
            "min": level['threshold'],
            "max": max_val,
            "color": level['color']
        }
    return result


def get_qualitative_grades_dict() -> dict:
    """Build a QUALITATIVE_GRADES-compatible dict from dynamic config."""
    config = get_score_levels_config()
    levels_by_key = {lvl['key']: lvl for lvl in config['levels']}
    result = {}

    for grade, level_key in config.get('qualitative_mapping', {}).items():
        level = levels_by_key.get(level_key)
        if level:
            result[grade] = {
                "score": level['threshold'],
                "level": level_key
            }
        else:
            # Fallback to lowest level
            lowest = sorted(config['levels'], key=lambda x: x['threshold'])[0]
            result[grade] = {
                "score": lowest['threshold'],
                "level": lowest['key']
            }
    return result


def get_level_name(level_key: str, lang: str = None) -> str:
    """Get translated level name from dynamic config."""
    if lang is None:
        lang = st.session_state.get('language', 'en')

    config = get_score_levels_config()
    for level in config['levels']:
        if level['key'] == level_key:
            return level['names'].get(lang, level['names'].get('en', level_key))
    return level_key


def validate_score_levels_config(config: dict) -> tuple:
    """Validate score levels configuration. Returns (is_valid, error_message)."""
    # Check minimum 2 levels
    if len(config.get('levels', [])) < 2:
        return False, "At least 2 levels are required"

    # Check unique thresholds
    thresholds = [lvl['threshold'] for lvl in config['levels']]
    if len(thresholds) != len(set(thresholds)):
        return False, "Thresholds must be unique"

    # Check non-empty names
    for lvl in config['levels']:
        if not lvl.get('names', {}).get('en', '').strip():
            return False, f"Level '{lvl['key']}' must have an English name"

    # Check min < max score
    if config['min_score'] >= config['max_score']:
        return False, "Min score must be less than max score"

    # Check all thresholds within range
    for lvl in config['levels']:
        if not (config['min_score'] <= lvl['threshold'] <= config['max_score']):
            return False, f"Threshold {lvl['threshold']} is outside score range"

    # Check qualitative mapping references valid levels
    level_keys = {lvl['key'] for lvl in config['levels']}
    for grade, level_key in config.get('qualitative_mapping', {}).items():
        if level_key not in level_keys:
            return False, f"Grade {grade} maps to unknown level '{level_key}'"

    return True, ""


def get_level_label(level_key: str) -> str:
    """Get translated label for a level key using dynamic config."""
    return get_level_name(level_key)


def calculate_score(actual, metric_type: str, thresholds: dict) -> dict:
    """Calculate score for a KR. Handles quantitative (higher/lower better) and qualitative (A/B/C/D/E) metrics.
    Uses dynamic score levels configuration.
    """
    config = get_score_levels_config()
    levels_dict = get_levels_dict()
    min_score = config['min_score']
    max_score = config['max_score']
    sorted_levels = sorted(config['levels'], key=lambda x: x['threshold'])

    # Handle qualitative metrics (A/B/C/D/E grades)
    if metric_type == "qualitative":
        qual_grades = get_qualitative_grades_dict()
        grade = str(actual).upper() if actual else "E"
        if grade in qual_grades:
            grade_info = qual_grades[grade]
            return {
                "score": grade_info["score"],
                "level": grade_info["level"],
                "level_info": levels_dict[grade_info["level"]],
                "grade": grade
            }
        else:
            # Default to lowest level if invalid grade
            lowest = sorted_levels[0]
            return {
                "score": lowest['threshold'],
                "level": lowest['key'],
                "level_info": levels_dict[lowest['key']],
                "grade": "E"
            }

    # Handle quantitative metrics
    actual = float(actual) if actual else 0.0

    # Only use levels that exist in the KR thresholds data
    # Filter out levels that don't have corresponding KR thresholds
    valid_levels = []
    for lvl in sorted_levels:
        if lvl['key'] in thresholds:
            valid_levels.append(lvl)

    # If no valid levels found, use the base 4 levels (below, meets, good, exceptional)
    if not valid_levels:
        base_keys = ['below', 'meets', 'good', 'exceptional']
        valid_levels = [lvl for lvl in sorted_levels if lvl['key'] in base_keys]

    # Still no valid levels? Use all sorted levels as fallback
    if not valid_levels:
        valid_levels = sorted_levels

    level_keys = [lvl['key'] for lvl in valid_levels]
    kr_thresholds = [thresholds.get(key, 0) for key in level_keys]
    level_scores = [lvl['threshold'] for lvl in valid_levels]

    score = min_score
    level_key = valid_levels[0]['key']

    if metric_type == "higher_better":
        # Find which level the actual value falls into (check from highest to lowest)
        for i in range(len(valid_levels) - 1, -1, -1):
            kr_th = kr_thresholds[i]
            if actual >= kr_th:
                if i == len(valid_levels) - 1:
                    # At or above top level
                    score = max_score
                else:
                    # Interpolate between this level and next
                    next_kr_th = kr_thresholds[i + 1]
                    if next_kr_th > kr_th:
                        ratio = (actual - kr_th) / (next_kr_th - kr_th)
                    else:
                        ratio = 1.0
                    base_score = level_scores[i]
                    next_score = level_scores[i + 1]
                    score = base_score + ratio * (next_score - base_score)

                level_key = valid_levels[i]['key']
                break
    else:
        # Lower is better - reverse logic
        for i in range(len(valid_levels) - 1, -1, -1):
            kr_th = kr_thresholds[i]
            if actual <= kr_th:
                if i == len(valid_levels) - 1:
                    score = max_score
                else:
                    next_kr_th = kr_thresholds[i + 1]
                    if next_kr_th < kr_th:
                        ratio = 1 - (actual - next_kr_th) / (kr_th - next_kr_th)
                    else:
                        ratio = 1.0
                    base_score = level_scores[i]
                    next_score = level_scores[i + 1]
                    score = base_score + ratio * (next_score - base_score)

                level_key = valid_levels[i]['key']
                break

    final_score = round(min(max(score, min_score), max_score), 2)
    return {"score": final_score, "level": level_key, "level_info": levels_dict[level_key]}


def get_level_for_score(score: float) -> dict:
    """Determine level for a given score using dynamic config."""
    config = get_score_levels_config()
    levels_dict = get_levels_dict()
    sorted_levels = sorted(config['levels'], key=lambda x: x['threshold'], reverse=True)

    for level in sorted_levels:
        if score >= level['threshold']:
            return {**levels_dict[level['key']], "key": level['key']}

    # Default to lowest level
    lowest = sorted(config['levels'], key=lambda x: x['threshold'])[0]
    return {**levels_dict[lowest['key']], "key": lowest['key']}


def score_to_percentage(score: float) -> float:
    """Convert score to percentage using dynamic config."""
    config = get_score_levels_config()
    min_score = config['min_score']
    max_score = config['max_score']
    score_range = max_score - min_score
    if score_range == 0:
        return 100.0 if score >= max_score else 0.0
    return round(((score - min_score) / score_range) * 100, 1)


def calculate_weighted_objective_score(objective: dict) -> dict:
    """
    Calculate weighted score for an Objective (weighted average of KR scores)
    Formula: OKR = (KR1 × weight1) + (KR2 × weight2) + ... / totalWeight
    Example: KR1(4.5 × 60%) + KR2(4.7 × 30%) + KR3(4.8 × 10%) = 4.59
    """
    krs = objective.get('key_results', [])
    if not krs:
        return {"score": 0, "level": get_level_for_score(0), "results": [], "formula_parts": [], "total_weight": 0}

    results = []
    formula_parts = []
    weighted_sum = 0
    total_weight = 0

    for kr in krs:
        result = calculate_score(kr['actual'], kr['metric_type'], kr.get('thresholds', {}))
        results.append(result)

        # Get KR weight (default to equal distribution if not set)
        kr_weight = kr.get('weight') or 0

        weighted_sum += result['score'] * kr_weight
        total_weight += kr_weight

        # Store formula part for display
        formula_parts.append({
            'name': kr['name'],
            'score': result['score'],
            'weight': kr_weight,
            'weighted_score': result['score'] * kr_weight / 100 if kr_weight > 0 else 0
        })

    # Calculate weighted average
    if total_weight > 0:
        avg_score = weighted_sum / total_weight
    else:
        # Fallback to simple average if no weights defined
        avg_score = sum(r['score'] for r in results) / len(results) if results else 0

    return {
        "score": round(avg_score, 2),
        "level": get_level_for_score(avg_score),
        "results": results,
        "formula_parts": formula_parts,
        "total_weight": total_weight
    }


def calculate_weighted_department_score(department: dict) -> dict:
    """
    Calculate weighted score for a Department based on objective weights.
    Formula: Dept = (Obj1 × weight1) + (Obj2 × weight2) + ... / totalWeight
    """
    objectives = department.get('objectives', [])
    if not objectives:
        return {"score": 0, "level": get_level_for_score(0), "objective_scores": [], "formula_parts": [],
                "total_weight": 0}

    # Count objectives with key results for default weight calculation
    objectives_with_krs = [obj for obj in objectives if obj.get('key_results')]

    if not objectives_with_krs:
        return {"score": 0, "level": get_level_for_score(0), "objective_scores": [], "formula_parts": [],
                "total_weight": 0}

    total_weight = 0
    weighted_sum = 0
    obj_scores = []
    formula_parts = []

    for obj in objectives:
        # Skip objectives with no key results
        if not obj.get('key_results'):
            continue

        # Get objective weight (default to equal distribution if not set)
        obj_weight = obj.get('weight') or (100.0 / len(objectives_with_krs))

        obj_result = calculate_weighted_objective_score(obj)
        obj_scores.append(obj_result)

        weighted_sum += obj_result['score'] * obj_weight
        total_weight += obj_weight

        # Store formula part for display
        formula_parts.append({
            'name': obj['name'],
            'score': obj_result['score'],
            'weight': obj_weight,
            'weighted_score': obj_result['score'] * obj_weight / 100 if obj_weight > 0 else 0
        })

    # Calculate weighted average (normalize if weights don't sum to 100)
    if total_weight > 0:
        avg_score = weighted_sum / total_weight
    else:
        avg_score = sum(s['score'] for s in obj_scores) / len(obj_scores) if obj_scores else 0

    return {
        "score": round(avg_score, 2),
        "level": get_level_for_score(avg_score),
        "objective_scores": obj_scores,
        "formula_parts": formula_parts,
        "total_weight": total_weight
    }


def create_gauge(score: float, compact: bool = False) -> str:
    """Returns HTML string with ECharts gauge using dynamic configuration."""
    import random
    config = get_score_levels_config()
    min_score = config['min_score']
    max_score = config['max_score']
    sorted_levels = sorted(config['levels'], key=lambda x: x['threshold'])
    score_range = max_score - min_score

    percentage = score_to_percentage(score)
    level_info = get_level_for_score(score)
    level_label = get_level_label(level_info['key'])

    # Use unique ID to avoid conflicts when multiple gauges on page
    gauge_id = f"gauge_{random.randint(10000, 99999)}"

    # Compact mode settings
    height = 180 if compact else 240
    font_size = 12 if compact else 16
    label_size = 8 if compact else 10
    pointer_width = 6 if compact else 10
    axis_width = 15 if compact else 24

    # Build dynamic color stops based on configured levels
    # ECharts color format: [position, color] where color applies UP TO that position
    color_stops = []
    for i, level in enumerate(sorted_levels):
        if i == len(sorted_levels) - 1:
            # Last level goes to 1.0
            stop_position = 1.0
        else:
            # Calculate position as fraction of total range (where this level ENDS)
            next_threshold = sorted_levels[i + 1]['threshold']
            stop_position = (next_threshold - min_score) / score_range if score_range > 0 else 1.0
        color_stops.append(f"[{stop_position:.3f}, '{level['color']}']")

    color_array = ",\n                            ".join(color_stops)

    # Build threshold values array for custom axis labels
    thresholds_js = [level['threshold'] for level in sorted_levels]
    thresholds_str = ", ".join([str(t) for t in thresholds_js])

    # Use splitNumber = 100 so we have positions at 0.00, 0.01, 0.02, ... 1.00
    # Then only show labels at positions matching our thresholds
    split_number = 100

    html = f'''
    <div id="{gauge_id}" style="width: 100%; height: {height}px;"></div>
    <script src="https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js"></script>
    <script>
        var chart = echarts.init(document.getElementById('{gauge_id}'));
        var thresholds = [{thresholds_str}];
        var option = {{
            series: [{{
                type: 'gauge',
                min: {min_score},
                max: {max_score},
                splitNumber: {split_number},
                radius: '90%',
                center: ['50%', '60%'],
                startAngle: 180,
                endAngle: 0,
                axisLine: {{
                    lineStyle: {{
                        width: {axis_width},
                        color: [
                            {color_array}
                        ]
                    }}
                }},
                pointer: {{
                    icon: 'path://M12.8,0.7l12,40.1H0.7L12.8,0.7z',
                    length: '60%',
                    width: {pointer_width},
                    offsetCenter: [0, '-10%'],
                    itemStyle: {{
                        color: '#1a1a2e'
                    }}
                }},
                axisTick: {{
                    show: false
                }},
                splitLine: {{
                    show: true,
                    length: 10,
                    distance: 0,
                    lineStyle: {{
                        color: '#444',
                        width: function(index) {{
                            return 0;
                        }}
                    }}
                }},
                axisLabel: {{
                    color: '#444',
                    fontSize: {label_size},
                    distance: -35,
                    formatter: function (value) {{
                        // Only show labels at threshold positions (with small tolerance for floating point)
                        for (var i = 0; i < thresholds.length; i++) {{
                            if (Math.abs(value - thresholds[i]) < 0.005) {{
                                return thresholds[i].toFixed(2);
                            }}
                        }}
                        return '';
                    }}
                }},
                title: {{
                    show: false
                }},
                detail: {{
                    fontSize: {font_size},
                    offsetCenter: [0, '55%'],
                    valueAnimation: true,
                    formatter: function (value) {{
                        return value.toFixed(2) + '\\n({percentage}%)';
                    }},
                    color: '{level_info['color']}',
                    fontFamily: 'Arial',
                    fontWeight: 'bold'
                }},
                data: [{{
                    value: {score},
                    name: '{level_label}',
                    itemStyle: {{
                        color: '{level_info['color']}'
                    }}
                }}]
            }}]
        }};
        chart.setOption(option);
    </script>
    '''
    return html


def render_sidebar(departments):
    """Render professional left sidebar with navigation and controls"""
    # Calculate overall stats using weighted calculations
    total_objectives = sum(len(d.get('objectives', [])) for d in departments)
    dept_scores = []

    for dept in departments:
        if dept.get('objectives'):
            dept_result = calculate_weighted_department_score(dept)
            dept_scores.append(dept_result['score'])

    avg_overall = round(sum(dept_scores) / len(dept_scores), 2) if dept_scores else 0
    overall_level = get_level_for_score(avg_overall) if dept_scores else {"color": THEME['text_secondary']}

    st.markdown(
        f"<h3 style='font-size:11px; color:{THEME['text_secondary']}; text-transform:uppercase; letter-spacing:1.5px; margin:0 0 16px 0; font-weight:600;'> {t('overview')}</h3>",
        unsafe_allow_html=True)

    # Stats cards with gradient backgrounds
    st.markdown(f"""
        <div style='background:linear-gradient(135deg, #e6f0ff 0%, #f0f7ff 100%); padding:16px; border-radius:10px; margin-bottom:12px; border:1px solid #cce0ff;'>
            <div style='font-size:32px; font-weight:700; color:#0066cc; line-height:1;'>{total_objectives}</div>
            <div style='font-size:11px; color:#4a90d9; font-weight:500; text-transform:uppercase; letter-spacing:0.5px; margin-top:4px;'>{t('total_objectives')}</div>
        </div>
    """, unsafe_allow_html=True)

    # Show department weighted average breakdown
    if dept_scores:
        st.markdown(f"""
            <div style='background:linear-gradient(135deg, {overall_level['color']}15 0%, {overall_level['color']}08 100%); padding:16px; border-radius:10px; margin-bottom:20px; border:1px solid {overall_level['color']}30;'>
                <div style='font-size:32px; font-weight:700; color:{overall_level['color']}; line-height:1;'>{avg_overall}</div>
                <div style='font-size:11px; color:{overall_level['color']}; font-weight:500; text-transform:uppercase; letter-spacing:0.5px; margin-top:4px;'>{t('dept_weighted_avg')}</div>
            </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown(f"""
            <div style='background:linear-gradient(135deg, {overall_level['color']}15 0%, {overall_level['color']}08 100%); padding:16px; border-radius:10px; margin-bottom:20px; border:1px solid {overall_level['color']}30;'>
                <div style='font-size:32px; font-weight:700; color:{overall_level['color']}; line-height:1;'>{avg_overall}</div>
                <div style='font-size:11px; color:{overall_level['color']}; font-weight:500; text-transform:uppercase; letter-spacing:0.5px; margin-top:4px;'>{t('weighted_score')}</div>
            </div>
        """, unsafe_allow_html=True)


def render_objective_card(objective, dept_idx, obj_idx, compact=True):
    """Render objective - grid view with compact cards OR full view with original detailed display"""
    krs = objective.get('key_results', [])
    if not krs:
        st.warning(t("no_krs"))
        return

    # Calculate weighted scores
    obj_result = calculate_weighted_objective_score(objective)
    avg_score = obj_result['score']
    results = obj_result['results']
    avg_level = get_level_for_score(avg_score)
    avg_pct = score_to_percentage(avg_score)
    obj_weight = objective.get('weight') or 0  # Objective weight within department (handles None)

    if compact:
        # GRID VIEW - Professional compact card with modern styling
        # Use Streamlit container with border for the card
        with st.container(border=True):
            # Header section - build badges section
            badges_html = f"""<span style='display:inline-block; padding:6px 14px; background:{avg_level['color']}15; color:{avg_level['color']}; border:1px solid {avg_level['color']}30; border-radius:8px; font-size:12px; font-weight:600; text-transform:uppercase; letter-spacing:0.5px;'>{get_level_label(avg_level['key'])} • {avg_pct}%</span>
                        <span style='display:inline-block; padding:6px 14px; background:#f1f5f9; color:{THEME['text_secondary']}; border:1px solid #e2e8f0; border-radius:8px; font-size:12px; font-weight:600;'>{len(krs)} KRs</span>"""

            if obj_weight > 0:
                badges_html += f"""
                        <span style='display:inline-block; padding:5px 12px; background:#fef3c7; color:#d97706; border-radius:6px; font-size:11px; font-weight:600;'>{t('weight')}: {obj_weight}%</span>"""

            st.markdown(f"""
                <div style='background:linear-gradient(180deg, {avg_level['color']}08 0%, #ffffff 100%); padding:16px; margin-bottom:16px; border-bottom:3px solid {avg_level['color']}; border-radius:8px 8px 0 0;'>
                    <div style='display:flex; justify-content:space-between; align-items:flex-start; gap:12px;'>
                        <h3 style='margin:0; font-size:16px; color:{THEME['text_primary']}; font-weight:700; flex:1; word-wrap:break-word; overflow-wrap:break-word; line-height:1.4;'>📋 {objective['name']}</h3>
                        <div style='background:linear-gradient(135deg, {avg_level['color']} 0%, {avg_level['color']}dd 100%); color:white; padding:8px 16px; border-radius:20px; font-size:15px; font-weight:700; white-space:nowrap; flex-shrink:0; box-shadow:0 3px 10px {avg_level['color']}50;'>{avg_score:.2f}</div>
                    </div>
                    <div style='margin-top:14px; display:flex; gap:10px; flex-wrap:wrap;'>
                        {badges_html}
                    </div>
                </div>
            """, unsafe_allow_html=True)

            gauge_html = create_gauge(avg_score, compact=False)
            components.html(gauge_html, height=260)

            # Editable table for facts with weight information
            table_data = []
            for kr_idx, kr in enumerate(krs):
                result = results[kr_idx]
                kr_weight = kr.get('weight', 0) or 0
                weighted_contribution = result['score'] * kr_weight / 100 if kr_weight > 0 else 0
                table_data.append({
                    "KR": f"KR{kr_idx + 1}",
                    t("key_result"): kr['name'],
                    t("weight"): f"{kr_weight}%",
                    t("fact"): kr['actual'],
                    "Score": result['score'],
                    "Weighted": round(weighted_contribution, 2),
                })

            df = pd.DataFrame(table_data)

            # Check if there are any qualitative metrics
            has_qualitative = any(kr['metric_type'] == 'qualitative' for kr in krs)

            # Editable table for Fact column
            # Use TextColumn if there are qualitative metrics, NumberColumn otherwise
            if has_qualitative:
                fact_column_config = st.column_config.TextColumn(t("fact"), width="small")
            else:
                fact_column_config = st.column_config.NumberColumn(t("fact"), min_value=-1000, max_value=10000,
                                                                   step=1, format="%.1f")

            edited_df = st.data_editor(
                df,
                column_config={
                    "KR": st.column_config.TextColumn("KR", disabled=True, width="small"),
                    t("key_result"): st.column_config.TextColumn(t("key_result"), disabled=True, width="medium"),
                    t("weight"): st.column_config.TextColumn(t("weight"), disabled=True, width="small"),
                    t("fact"): fact_column_config,
                    "Score": st.column_config.NumberColumn("Score", disabled=True, format="%.2f", width="small"),
                    "Weighted": st.column_config.NumberColumn("Weighted", disabled=True, format="%.2f", width="small"),
                },
                hide_index=True,
                use_container_width=True,
                key=f"grid_editor_d{dept_idx}_o{obj_idx}_{objective['id']}"
            )

            # Update actual values from edited dataframe
            for i, row in edited_df.iterrows():
                if i < len(krs):
                    new_actual = row[t("fact")]
                    if new_actual != krs[i]['actual']:
                        st.session_state.departments[dept_idx]['objectives'][obj_idx]['key_results'][i][
                            'actual'] = new_actual
                        save_data()
                        st.rerun()

            # Edit Objective in Grid view
            with st.expander(f"{t('edit_objective')}", expanded=False):
                grid_obj_col1, grid_obj_col2 = st.columns([3, 1])
                with grid_obj_col1:
                    grid_edit_obj_name = st.text_input(
                        t("objective_name"),
                        value=objective['name'],
                        key=f"grid_edit_obj_name_d{dept_idx}_o{obj_idx}_{objective['id']}"
                    )
                with grid_obj_col2:
                    grid_edit_obj_weight = st.number_input(
                        t("objective_weight"),
                        value=float(objective.get('weight', 0) or 0),
                        min_value=0.0,
                        max_value=100.0,
                        step=1.0,
                        key=f"grid_edit_obj_weight_d{dept_idx}_o{obj_idx}_{objective['id']}"
                    )

                if st.button(f"✅ {t('update')}", key=f"grid_update_obj_btn_d{dept_idx}_o{obj_idx}_{objective['id']}"):
                    if grid_edit_obj_name.strip():
                        st.session_state.departments[dept_idx]['objectives'][obj_idx][
                            'name'] = grid_edit_obj_name.strip()
                        st.session_state.departments[dept_idx]['objectives'][obj_idx]['weight'] = grid_edit_obj_weight
                        save_data()
                        st.rerun()

            with st.expander(f"{t('delete_krs')}", expanded=False):
                for kr_idx, kr in enumerate(krs):
                    if st.button(f"{t('delete')} KR{kr_idx + 1}", key=f"del_grid_kr_d{dept_idx}_o{obj_idx}_{kr['id']}"):
                        st.session_state.departments[dept_idx]['objectives'][obj_idx]['key_results'] = [
                            k for k in krs if k['id'] != kr['id']
                        ]
                        save_data()
                        st.rerun()

                if st.button(f"🗑️ {t('delete_objective')}", key=f"del_obj_d{dept_idx}_o{obj_idx}", type="secondary"):
                    st.session_state.departments[dept_idx]['objectives'] = [
                        o for o in st.session_state.departments[dept_idx]['objectives'] if o['id'] != objective['id']
                    ]
                    save_data()
                    st.rerun()

    else:
        # FULL VIEW - Original detailed display with all tables and functionality wrapped in frame
        obj_weight = objective.get('weight') or 0  # handles None values
        weight_badge = f"<span style='background:#fef3c7; color:#d97706; padding:4px 10px; border-radius:12px; font-weight:600; font-size:12px; margin-left:8px;'>{t('weight')}: {obj_weight}%</span>" if obj_weight > 0 else ""

        st.markdown(
            f"<div style='background:{THEME['card_bg']}; border:none; border-radius:10px; padding:0; margin-bottom:20px; box-shadow:0 4px 12px rgba(0,0,0,0.1); overflow:hidden;'>",
            unsafe_allow_html=True)
        st.markdown(
            f"<div style='background:#FFC000; padding:8px 12px; border-radius:5px; display:flex; justify-content:space-between; align-items:center; margin-bottom:10px;'><span style='font-weight:bold; font-size:14px;'>📋 {objective['name']}{weight_badge}</span><span style='background:{avg_level['color']}; color:white; padding:4px 12px; border-radius:15px; font-weight:bold; font-size:14px;'>{t('weighted_score')}: {avg_score:.2f}</span></div>",
            unsafe_allow_html=True)

        with st.expander(f"{objective['name']}", expanded=False):
            col_table, col_gauge = st.columns([2, 1])

            with col_table:
                # Build DataFrame for editable table with weight information
                table_data = []
                for kr_idx, kr in enumerate(krs):
                    result = results[kr_idx]
                    kr_weight = kr.get('weight', 0) or 0
                    weighted_contribution = result['score'] * kr_weight / 100 if kr_weight > 0 else 0
                    table_data.append({
                        "KR": f"KR{kr_idx + 1}",
                        t("key_result"): kr['name'],
                        t("weight"): f"{kr_weight}%",
                        t("fact"): kr['actual'],
                        "Score": result['score'],
                        "Weighted": round(weighted_contribution, 2),
                    })

                df = pd.DataFrame(table_data)

                # Check if there are any qualitative metrics
                has_qualitative = any(kr['metric_type'] == 'qualitative' for kr in krs)

                # Editable table for Fact column
                # Use TextColumn if there are qualitative metrics, NumberColumn otherwise
                if has_qualitative:
                    fact_column_config = st.column_config.TextColumn(t("fact"), width="small")
                else:
                    fact_column_config = st.column_config.NumberColumn(t("fact"), min_value=-1000, max_value=10000,
                                                                       step=1, format="%.1f")

                edited_df = st.data_editor(
                    df,
                    column_config={
                        "KR": st.column_config.TextColumn("KR", disabled=True, width="small"),
                        t("key_result"): st.column_config.TextColumn(t("key_result"), disabled=True,
                                                                     width="medium"),
                        t("weight"): st.column_config.TextColumn(t("weight"), disabled=True, width="small"),
                        t("fact"): fact_column_config,
                        "Score": st.column_config.NumberColumn("Score", disabled=True, format="%.2f",
                                                               width="small"),
                        "Weighted": st.column_config.NumberColumn("Weighted", disabled=True, format="%.2f",
                                                                  width="small"),
                    },
                    hide_index=True,
                    use_container_width=True,
                    key=f"editor_d{dept_idx}_o{obj_idx}_{objective['id']}"
                )

                # Update actual values from edited dataframe
                for i, row in edited_df.iterrows():
                    if i < len(krs):
                        new_actual = row[t("fact")]
                        if new_actual != krs[i]['actual']:
                            st.session_state.departments[dept_idx]['objectives'][obj_idx]['key_results'][i][
                                'actual'] = new_actual
                            save_data()
                            st.rerun()

                # Results breakdown table (without weights)
                st.markdown(f"#### {t('results_breakdown')}")

                # Get dynamic config levels - show ALL configured levels
                score_config = get_score_levels_config()
                sorted_config_levels = sorted(score_config['levels'], key=lambda x: x['threshold'])

                # Use all configured levels for display
                display_levels = sorted_config_levels

                # Build dynamic header for levels
                level_headers = ""
                for lvl in display_levels:
                    text_color = "#000" if lvl['key'] == 'meets' else "white"
                    level_headers += f"<th style='padding:6px; border:1px solid #2F5496; background:{lvl['color']}; color:{text_color}; font-size:10px;'>{get_level_label(lvl['key'])}<br><small style='font-size:9px;'>{lvl['threshold']:.2f}</small></th>"

                num_level_cols = len(display_levels)
                total_cols = 4 + num_level_cols + 1  # KR, name, weight, fact + level cols + result

                html_table = f"<table style='width:100%; border-collapse:collapse; font-size:11px; margin-top:5px;'><thead><tr style='background:#4472C4; color:white;'><th style='padding:6px; border:1px solid #2F5496; font-size:11px;'>KR</th><th style='padding:6px; border:1px solid #2F5496; font-size:11px;'>{t('key_result')}</th><th style='padding:6px; border:1px solid #2F5496; font-size:11px;'>{t('weight')}</th><th style='padding:6px; border:1px solid #2F5496; font-size:11px;'>{t('fact')}</th>{level_headers}<th style='padding:6px; border:1px solid #2F5496; font-size:11px;'>{t('result')}</th></tr></thead><tbody>"

                for kr_idx, kr in enumerate(krs):
                    result = results[kr_idx]
                    th = kr.get('thresholds', {})
                    level = result['level']

                    # Build cells dictionary dynamically based on display_levels
                    cells = {}
                    for lvl in display_levels:
                        if level == lvl['key']:
                            text_color = "#000" if lvl['key'] == 'meets' else "white"
                            cells[lvl['key']] = f"background:{lvl['color']}; color:{text_color}; font-weight:bold;"
                        else:
                            cells[lvl['key']] = ''

                    # Handle qualitative vs quantitative display
                    if kr['metric_type'] == 'qualitative':
                        actual_display = kr.get('actual', 'E')
                        th_texts = {lvl['key']: chr(ord('E') - i) for i, lvl in enumerate(display_levels)}
                    elif kr['metric_type'] == "higher_better":
                        actual_display = f"{kr['actual']}{kr.get('unit', '')}"
                        th_texts = {}
                        for i, lvl in enumerate(display_levels):
                            if i == 0:
                                th_texts[lvl['key']] = f"<{th.get(lvl['key'], 0)}"
                            else:
                                th_texts[lvl['key']] = f"≥{th.get(lvl['key'], 0)}"
                    else:
                        actual_display = f"{kr['actual']}{kr.get('unit', '')}"
                        th_texts = {}
                        for i, lvl in enumerate(display_levels):
                            if i == 0:
                                th_texts[lvl['key']] = f">{th.get(lvl['key'], 0)}"
                            else:
                                th_texts[lvl['key']] = f"≤{th.get(lvl['key'], 0)}"

                    row_bg = '#F8F9FA' if kr_idx % 2 == 0 else '#FFFFFF'
                    kr_desc = kr.get('description', '') or kr['name']
                    kr_desc_escaped = kr_desc.replace('"', '&quot;').replace("'", "&#39;")

                    # Get KR weight and calculate weighted contribution
                    kr_weight = kr.get('weight', 0) or 0
                    weighted_contribution = result['score'] * kr_weight / 100 if kr_weight > 0 else 0

                    # Show score with weight contribution in result cell
                    if kr_weight > 0:
                        score_display = f"{result['score']:.2f}<br><small style='font-size:9px;'>x {kr_weight}% = {weighted_contribution:.2f}</small>"
                    else:
                        score_display = f"{result['score']:.2f}"

                    # Build level cells dynamically
                    level_cells = ""
                    for lvl in display_levels:
                        level_cells += f"<td style='padding:5px; border:1px solid #ddd; {cells[lvl['key']]} font-size:11px;'>{th_texts[lvl['key']]}</td>"

                    html_table += f"<tr style='background:{row_bg};'><td style='padding:5px; border:1px solid #ddd; font-weight:bold; font-size:11px;'>KR{kr_idx + 1}</td><td style='padding:5px; border:1px solid #ddd; text-align:left; font-size:11px;' title=\"{kr_desc_escaped}\"><span style='cursor:help; border-bottom:1px dotted #7f8c8d;'>{kr['name']}</span></td><td style='padding:5px; border:1px solid #ddd; background:#FFF2CC; font-weight:bold; font-size:11px;'>{kr_weight}%</td><td style='padding:5px; border:1px solid #ddd; background:#E2EFDA; font-weight:bold; font-size:11px;'>{actual_display}</td>{level_cells}<td style='padding:5px; border:1px solid #ddd; background:{result['level_info']['color']}; color:white; font-weight:bold; font-size:11px;'>{score_display}</td></tr>"

                # Weighted Calculation Row - matches Java frontend format
                if obj_result.get('total_weight', 0) > 0:
                    # Build weighted formula string: OKR = (KR1 x w1%) + (KR2 x w2%) + ...
                    formula_parts = []
                    for kr_idx, kr in enumerate(krs):
                        kr_score = results[kr_idx]['score']
                        kr_weight = kr.get('weight', 0) or 0
                        formula_parts.append(f"({kr_score:.2f} x {kr_weight}%)")

                    formula_str = " + ".join(formula_parts)

                    # Calculate intermediate sum
                    weighted_contributions = sum(
                        (results[i]['score'] * (krs[i].get('weight', 0) or 0) / 100) for i in range(len(krs)))

                    # Show normalization if weights don't sum to 100
                    total_weight = sum((kr.get('weight', 0) or 0) for kr in krs)
                    if total_weight != 100 and total_weight > 0:
                        html_table += f"<tr style='background:#FFF2CC; font-weight:bold;'><td colspan='{total_cols - 1}' style='padding:8px; border:2px solid #BF9000; text-align:right; font-size:11px;'><span style='font-weight:bold;'>OKR =</span> {formula_str} = {weighted_contributions:.2f} / {total_weight}% = </td><td style='padding:8px; border:2px solid #BF9000; background:{avg_level['color']}; color:white; font-size:14px;'>{avg_score:.2f}</td></tr></tbody></table>"
                    else:
                        html_table += f"<tr style='background:#FFF2CC; font-weight:bold;'><td colspan='{total_cols - 1}' style='padding:8px; border:2px solid #BF9000; text-align:right; font-size:11px;'><span style='font-weight:bold;'>OKR =</span> {formula_str} = </td><td style='padding:8px; border:2px solid #BF9000; background:{avg_level['color']}; color:white; font-size:14px;'>{avg_score:.2f}</td></tr></tbody></table>"
                else:
                    # Fallback to simple average formula (no weights)
                    kr_formula = " + ".join([f"KR{i + 1}" for i in range(len(krs))])
                    html_table += f"<tr style='background:#FFF2CC; font-weight:bold;'><td colspan='{total_cols - 1}' style='padding:8px; border:2px solid #BF9000; text-align:right; font-size:11px;'>({kr_formula}) / {len(krs)} =</td><td style='padding:8px; border:2px solid #BF9000; background:{avg_level['color']}; color:white; font-size:14px;'>{avg_score:.2f}</td></tr></tbody></table>"

                table_height = 70 + (len(krs) * 48) + 60
                components.html(html_table, height=table_height, scrolling=False)

                # Edit KR section
                st.markdown(f"#### {t('edit_krs')}")
                for kr_idx, kr in enumerate(krs):
                    with st.expander(f"✏️ {t('edit_kr')} {kr_idx + 1}: {kr['name']}", expanded=False):
                        ec1, ec2, ec3 = st.columns([3, 2, 1])
                        with ec1:
                            edit_name = st.text_input(t("kr_name"), value=kr['name'],
                                                      key=f"edit_name_d{dept_idx}_o{obj_idx}_kr{kr_idx}_{kr['id']}")
                        with ec2:
                            # Determine current metric type
                            current_type_idx = 0 if kr['metric_type'] == "higher_better" else 1 if kr[
                                                                                                       'metric_type'] == "lower_better" else 2
                            edit_type = st.selectbox(
                                t("type"),
                                ["higher_better", "lower_better", "qualitative"],
                                index=current_type_idx,
                                format_func=lambda
                                    x: "↑ Higher is better" if x == "higher_better" else "↓ Lower is better" if x == "lower_better" else "⭐ Qualitative (A/B/C/D/E)",
                                key=f"edit_type_d{dept_idx}_o{obj_idx}_kr{kr_idx}_{kr['id']}"
                            )
                        with ec3:
                            edit_unit = st.text_input(t("unit"), value=kr.get('unit', '%'),
                                                      key=f"edit_unit_d{dept_idx}_o{obj_idx}_kr{kr_idx}_{kr['id']}")

                        # Description field
                        edit_description = st.text_area(
                            t("kr_description"),
                            value=kr.get('description', ''),
                            placeholder=t("kr_description_placeholder"),
                            key=f"edit_desc_d{dept_idx}_o{obj_idx}_kr{kr_idx}_{kr['id']}",
                            height=68
                        )

                        # Weight field
                        edit_weight = st.number_input(
                            t("kr_weight"),
                            value=float(kr.get('weight', 0) or 0),
                            min_value=0.0,
                            max_value=100.0,
                            step=1.0,
                            key=f"edit_weight_d{dept_idx}_o{obj_idx}_kr{kr_idx}_{kr['id']}"
                        )

                        # Thresholds - only show for quantitative metrics (dynamic levels)
                        edit_level_config = get_score_levels_config()
                        edit_sorted_lvls = sorted(edit_level_config['levels'], key=lambda x: x['order'])
                        th = kr.get('thresholds', {})

                        if edit_type != "qualitative":
                            st.markdown("**Thresholds**")
                            edit_threshold_cols = st.columns(len(edit_sorted_lvls))
                            edit_threshold_values = {}

                            for idx, lvl in enumerate(edit_sorted_lvls):
                                default_val = (100.0 / (len(edit_sorted_lvls) - 1)) * idx if len(edit_sorted_lvls) > 1 else 0.0
                                with edit_threshold_cols[idx]:
                                    st.markdown(f"<small style='color:{lvl['color']};'>● {lvl['threshold']:.2f}</small>", unsafe_allow_html=True)
                                    edit_threshold_values[lvl['key']] = st.number_input(
                                        get_level_label(lvl['key']),
                                        value=float(th.get(lvl['key'], default_val)),
                                        key=f"edit_{lvl['key']}_d{dept_idx}_o{obj_idx}_kr{kr_idx}_{kr['id']}"
                                    )
                        else:
                            # For qualitative, set default thresholds (not used but needed for data structure)
                            edit_threshold_values = {lvl['key']: 0.0 for lvl in edit_sorted_lvls}

                        if st.button(f"✅ {t('update')} KR{kr_idx + 1}",
                                     key=f"update_btn_d{dept_idx}_o{obj_idx}_kr{kr_idx}_{kr['id']}"):
                            if edit_name.strip():
                                st.session_state.departments[dept_idx]['objectives'][obj_idx]['key_results'][
                                    kr_idx].update({
                                    "name": edit_name.strip(),
                                    "metric_type": edit_type,
                                    "unit": edit_unit,
                                    "description": edit_description.strip(),
                                    "weight": edit_weight,
                                    "thresholds": edit_threshold_values
                                })
                                save_data()
                                st.rerun()

                # Edit Objective section
                with st.expander(f"{t('edit_objective')}: {objective['name']}", expanded=False):
                    obj_col1, obj_col2 = st.columns([3, 1])
                    with obj_col1:
                        edit_obj_name = st.text_input(
                            t("objective_name"),
                            value=objective['name'],
                            key=f"edit_obj_name_d{dept_idx}_o{obj_idx}_{objective['id']}"
                        )
                    with obj_col2:
                        edit_obj_weight = st.number_input(
                            t("objective_weight"),
                            value=float(objective.get('weight', 0) or 0),
                            min_value=0.0,
                            max_value=100.0,
                            step=1.0,
                            key=f"edit_obj_weight_d{dept_idx}_o{obj_idx}_{objective['id']}"
                        )

                    if st.button(f"✅ {t('update')} {t('objective')}",
                                 key=f"update_obj_btn_d{dept_idx}_o{obj_idx}_{objective['id']}"):
                        if edit_obj_name.strip():
                            st.session_state.departments[dept_idx]['objectives'][obj_idx][
                                'name'] = edit_obj_name.strip()
                            st.session_state.departments[dept_idx]['objectives'][obj_idx]['weight'] = edit_obj_weight
                            save_data()
                            st.rerun()

                st.markdown(f"#### {t('delete_krs')}")
                del_cols = st.columns(len(krs) + 1)
                for kr_idx, kr in enumerate(krs):
                    with del_cols[kr_idx]:
                        if st.button(f"{t('delete')} KR{kr_idx + 1}", key=f"del_kr_d{dept_idx}_o{obj_idx}_{kr['id']}"):
                            st.session_state.departments[dept_idx]['objectives'][obj_idx]['key_results'] = [k for k in
                                                                                                            krs if
                                                                                                            k['id'] !=
                                                                                                            kr['id']]
                            save_data()
                            st.rerun()

            with col_gauge:
                st.markdown(f"### {t('score')}")
                gauge_html = create_gauge(avg_score)
                components.html(gauge_html, height=260)

                st.markdown(
                    f"<div style='text-align:center; margin-top:8px;'><div style='background:{avg_level['color']}; color:white; padding:10px; border-radius:8px; font-size:16px; font-weight:bold;'>{get_level_label(avg_level['key'])}<br><small style='font-size:14px;'>{avg_score:.2f} ({avg_pct}%)</small></div></div>",
                    unsafe_allow_html=True)

            # Add KR section
            with st.expander(t("add_kr_to_obj")):
                ac1, ac2, ac3 = st.columns([3, 2, 1])
                with ac1:
                    add_name = st.text_input(t("kr_name"), key=f"add_name_d{dept_idx}_o{obj_idx}")
                with ac2:
                    add_type = st.selectbox(t("type"), ["higher_better", "lower_better"],
                                            format_func=lambda x: "↑" if x == "higher_better" else "↓",
                                            key=f"add_type_d{dept_idx}_o{obj_idx}")
                with ac3:
                    add_unit = st.text_input(t("unit"), value="%", key=f"add_unit_d{dept_idx}_o{obj_idx}")

                # Description field for tooltip
                add_description = st.text_area(t("kr_description"), placeholder=t("kr_description_placeholder"),
                                               key=f"add_desc_d{dept_idx}_o{obj_idx}", height=68)

                # Weight field
                add_weight = st.number_input(
                    t("kr_weight"),
                    value=0.0,
                    min_value=0.0,
                    max_value=100.0,
                    step=1.0,
                    key=f"add_weight_d{dept_idx}_o{obj_idx}"
                )

                # Dynamic threshold inputs based on config
                add_level_config = get_score_levels_config()
                add_sorted_lvls = sorted(add_level_config['levels'], key=lambda x: x['order'])
                add_threshold_cols = st.columns(len(add_sorted_lvls))
                add_threshold_values = {}

                for idx, lvl in enumerate(add_sorted_lvls):
                    default_val = (100.0 / (len(add_sorted_lvls) - 1)) * idx if len(add_sorted_lvls) > 1 else 0.0
                    with add_threshold_cols[idx]:
                        st.markdown(f"<small style='color:{lvl['color']};'>● {lvl['threshold']:.2f}</small>", unsafe_allow_html=True)
                        add_threshold_values[lvl['key']] = st.number_input(
                            get_level_label(lvl['key']),
                            value=round(default_val, 1),
                            key=f"add_{lvl['key']}_d{dept_idx}_o{obj_idx}"
                        )

                if st.button(t("add"), key=f"add_btn_d{dept_idx}_o{obj_idx}"):
                    if add_name.strip():
                        st.session_state.departments[dept_idx]['objectives'][obj_idx]['key_results'].append({
                            "id": str(uuid.uuid4()), "name": add_name.strip(), "metric_type": add_type,
                            "unit": add_unit, "description": add_description.strip(),
                            "weight": add_weight,
                            "thresholds": add_threshold_values,
                            "actual": 0.0
                        })
                        save_data()
                        st.rerun()

            # Delete objective
            if st.button(f"{t('delete_objective')} '{objective['name']}'",
                         key=f"del_obj_d{dept_idx}_{objective['id']}"):
                st.session_state.departments[dept_idx]['objectives'] = [o for o in
                                                                        st.session_state.departments[dept_idx][
                                                                            'objectives'] if o['id'] != objective['id']]
                save_data()
                st.rerun()

        # Close the frame container for full view
        st.markdown("</div>", unsafe_allow_html=True)


def save_data():
    data = {
        "departments": st.session_state.departments,
        "language": st.session_state.get('language', 'en'),
        "score_levels_config": get_score_levels_config()
    }
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)

            # Migration: convert old flat structure to department structure
            if 'objectives' in data and 'departments' not in data:
                # Old format: migrate to new format
                old_objectives = data.get('objectives', [])
                if old_objectives:
                    # Create a default department with all existing objectives
                    departments = [{
                        "id": str(uuid.uuid4()),
                        "name": "Default Department",
                        "objectives": old_objectives
                    }]
                else:
                    departments = []
                # Return with default config for old format
                return departments, data.get('language', 'en'), DEFAULT_SCORE_LEVELS_CONFIG
            else:
                # New format - load config or use default
                score_config = data.get('score_levels_config', DEFAULT_SCORE_LEVELS_CONFIG)
                return data.get('departments', []), data.get('language', 'en'), score_config
    return [], 'en', DEFAULT_SCORE_LEVELS_CONFIG


def _create_score_formula(row: int, metric_type: str, num_levels: int = 4) -> str:
    """Create Excel formula for score calculation based on metric type - using dynamic config."""
    config = get_score_levels_config()
    sorted_levels = sorted(config['levels'], key=lambda x: x['threshold'])
    min_score = config['min_score']
    max_score = config['max_score']

    # Column references - threshold columns start at I (after KR weight column)
    actual_col = 'G'
    base_threshold_col = ord('I')

    actual = f'{actual_col}{row}'

    # Build threshold column references for each level
    threshold_refs = {}
    for i, lvl in enumerate(sorted_levels):
        col_letter = chr(base_threshold_col + i)
        threshold_refs[lvl['key']] = f'{col_letter}{row}'

    # Build qualitative formula based on grade mapping
    qual_mapping = config.get('qualitative_mapping', {})
    qual_parts = []
    for grade in ['A', 'B', 'C', 'D', 'E']:
        if grade in qual_mapping:
            level_key = qual_mapping[grade]
            for lvl in sorted_levels:
                if lvl['key'] == level_key:
                    qual_parts.append(f'IF({actual}="{grade}",{lvl["threshold"]}')
                    break
    if qual_parts:
        qualitative_formula = ','.join(qual_parts) + f',{min_score}' + ')' * len(qual_parts)
    else:
        qualitative_formula = str(min_score)

    # Build higher_better formula (nested IFs for N levels)
    higher_parts = []
    for i in range(len(sorted_levels) - 1, -1, -1):
        lvl = sorted_levels[i]
        th = threshold_refs[lvl['key']]
        score = lvl['threshold']

        if i == len(sorted_levels) - 1:
            # Top level
            higher_parts.append(f'IF({actual}>={th},{max_score}')
        else:
            next_lvl = sorted_levels[i + 1]
            next_th = threshold_refs[next_lvl['key']]
            next_score = next_lvl['threshold']
            step = next_score - score
            higher_parts.append(f'IF({actual}>={th},{score}+({actual}-{th})/MAX({next_th}-{th},1)*{step}')

    higher_better_formula = ','.join(higher_parts) + f',{min_score}' + ')' * len(higher_parts)

    # Build lower_better formula (similar logic, reversed)
    lower_parts = []
    for i in range(len(sorted_levels) - 1, -1, -1):
        lvl = sorted_levels[i]
        th = threshold_refs[lvl['key']]
        score = lvl['threshold']

        if i == len(sorted_levels) - 1:
            lower_parts.append(f'IF({actual}<={th},{max_score}')
        else:
            next_lvl = sorted_levels[i + 1]
            next_th = threshold_refs[next_lvl['key']]
            next_score = next_lvl['threshold']
            step = next_score - score
            lower_parts.append(f'IF({actual}<={th},{score}+(1-({actual}-{th})/MAX({next_th}-{th},1))*{step}')

    lower_better_formula = ','.join(lower_parts) + f',{min_score}' + ')' * len(lower_parts)

    # Main formula that checks metric type
    if metric_type == 'qualitative':
        return f'={qualitative_formula}'
    elif metric_type == 'higher_better':
        return f'={higher_better_formula}'
    else:  # lower_better
        return f'={lower_better_formula}'


def _create_performance_level_formula(row: int, num_levels: int = 4) -> str:
    """Create Excel formula for performance level categorization - using dynamic config."""
    config = get_score_levels_config()
    sorted_levels = sorted(config['levels'], key=lambda x: x['threshold'], reverse=True)

    # Calculate score column based on number of levels (base columns + N threshold columns + score column)
    # Columns: A=Dept, B=Obj, C=ObjWeight, D=KR, E=KRWeight, F=Type, G=Actual, H=Unit, I+N-1=Thresholds, then Score
    score_col_idx = 8 + num_levels + 1  # After thresholds
    score_col = chr(ord('A') + score_col_idx - 1)
    score_cell = f'{score_col}{row}'

    # Build nested IF formula for N levels
    parts = []
    for lvl in sorted_levels:
        label = get_level_name(lvl['key'], 'en')  # Use English for Excel
        parts.append(f'IF({score_cell}>={lvl["threshold"]},"{label}"')

    # Default to lowest level name
    lowest_label = get_level_name(sorted_levels[-1]['key'], 'en')
    formula = ','.join(parts) + f',"{lowest_label}"' + ')' * len(parts)

    return f'={formula}'


def _apply_conditional_formatting(ws, max_row: int, colors: dict, num_levels: int = 4):
    """Apply conditional formatting to score and performance level columns - using dynamic config."""
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles.differential import DifferentialStyle

    # Skip if no data rows (max_row must be at least 3 to have row 2 as data)
    if max_row < 3:
        return

    config = get_score_levels_config()
    sorted_levels = sorted(config['levels'], key=lambda x: x['threshold'], reverse=True)

    white_font = Font(bold=True, color='FFFFFF')

    # Calculate score column based on number of levels
    # Columns: A=Dept, B=Obj, C=ObjWeight, D=KR, E=KRWeight, F=Type, G=Actual, H=Unit, I+N-1=Thresholds, then Score, Level
    score_col_idx = 8 + num_levels + 1
    level_col_idx = score_col_idx + 1
    score_col = chr(ord('A') + score_col_idx - 1)
    level_col = chr(ord('A') + level_col_idx - 1)

    score_range = f'{score_col}2:{score_col}{max_row - 1}'
    level_range = f'{level_col}2:{level_col}{max_row - 1}'

    # Apply conditional formatting rules for each level
    for priority, lvl in enumerate(sorted_levels, start=1):
        color_hex = lvl['color'].lstrip('#')
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')

        threshold = lvl['threshold']

        # Score column rule
        if priority == 1:
            # Top level: >= threshold
            rule = Rule(type='cellIs', operator='greaterThanOrEqual',
                       formula=[str(threshold)],
                       stopIfTrue=True,
                       dxf=DifferentialStyle(fill=fill, font=white_font))
        else:
            # Other levels: between this and previous threshold
            prev_lvl = sorted_levels[priority - 2]
            upper_bound = round(prev_lvl['threshold'] - 0.01, 2)
            rule = Rule(type='cellIs', operator='between',
                       formula=[str(threshold), str(upper_bound)],
                       stopIfTrue=True,
                       dxf=DifferentialStyle(fill=fill, font=white_font))

        rule.priority = priority
        ws.conditional_formatting.add(score_range, rule)

        # Performance level column rule (based on score column)
        if priority == 1:
            lrule = Rule(type='expression', formula=[f'${score_col}2>={threshold}'],
                        stopIfTrue=True,
                        dxf=DifferentialStyle(fill=fill, font=white_font))
        else:
            prev_lvl = sorted_levels[priority - 2]
            lrule = Rule(type='expression',
                        formula=[f'AND(${score_col}2>={threshold},${score_col}2<{prev_lvl["threshold"]})'],
                        stopIfTrue=True,
                        dxf=DifferentialStyle(fill=fill, font=white_font))

        lrule.priority = priority
        ws.conditional_formatting.add(level_range, lrule)


def export_to_excel(departments):
    """Export OKR data to Excel with interactive formulas for automatic recalculation"""
    wb = Workbook()
    ws = wb.active
    ws.title = "OKR Export"

    # Get dynamic level configuration
    config = get_score_levels_config()
    sorted_levels = sorted(config['levels'], key=lambda x: x['order'])
    num_levels = len(sorted_levels)

    # Build colors dict from config
    colors = {lvl['key']: lvl['color'].lstrip('#') for lvl in sorted_levels}

    # Define header style
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    weight_fill = PatternFill(start_color='d97706', end_color='d97706', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Build dynamic headers with level names
    level_headers = [get_level_name(lvl['key'], 'en') for lvl in sorted_levels]
    headers = [t('department'), t('objective'), t('objective_weight'), t('key_result'),
               t('kr_weight'), t('type'), t('actual'), t('unit')] + level_headers + [t('score').replace('🎯 ', ''), t('performance_level')]

    # Calculate column indices
    THRESHOLD_START_COL = 9  # Column I (shifted by 1 for KR weight)
    SCORE_COL = THRESHOLD_START_COL + num_levels
    LEVEL_COL = SCORE_COL + 1
    TOTAL_COLS = LEVEL_COL

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = weight_fill if header in [t('objective_weight'), t('kr_weight')] else header_fill
        cell.alignment = header_alignment

    # Add data
    row_idx = 2
    for dept in departments:
        dept_name = dept['name']
        dept_start_row = row_idx  # Track starting row for this department

        for obj in dept.get('objectives', []):
            obj_name = obj['name']
            obj_weight = obj.get('weight') or 0  # handles None values
            obj_start_row = row_idx  # Track starting row for this objective
            kr_list = obj.get('key_results', [])

            for kr in kr_list:
                # Calculate initial score using Python (for fallback coloring)
                result = calculate_score(kr['actual'], kr['metric_type'], kr.get('thresholds', {}))
                initial_score = result['score']
                initial_level = result['level']

                # Determine metric type display
                if kr['metric_type'] == 'qualitative':
                    type_display = 'Qualitative (A-E)'
                    actual_display = kr.get('actual', 'E')
                elif kr['metric_type'] == 'higher_better':
                    type_display = '↑ Higher Better'
                    actual_display = kr['actual']
                else:
                    type_display = '↓ Lower Better'
                    actual_display = kr['actual']

                # Write data (department only in first row of dept, objective only in first row of obj)
                kr_weight = kr.get('weight') or 0  # handles None values
                ws.cell(row=row_idx, column=1, value=dept_name if row_idx == dept_start_row else '')
                ws.cell(row=row_idx, column=2, value=obj_name if row_idx == obj_start_row else '')
                ws.cell(row=row_idx, column=3, value=f"{obj_weight}%" if row_idx == obj_start_row else '')
                ws.cell(row=row_idx, column=4, value=kr['name'])
                ws.cell(row=row_idx, column=5, value=f"{kr_weight}%")
                ws.cell(row=row_idx, column=6, value=type_display)
                ws.cell(row=row_idx, column=7, value=actual_display)
                ws.cell(row=row_idx, column=8, value=kr.get('unit', ''))

                # Dynamic thresholds based on levels
                th = kr.get('thresholds', {})
                if kr['metric_type'] == 'qualitative':
                    # For qualitative, show grade letters
                    qual_grades = ['E', 'D', 'C', 'B', 'A']
                    for i, lvl in enumerate(sorted_levels):
                        grade = qual_grades[i] if i < len(qual_grades) else ''
                        ws.cell(row=row_idx, column=THRESHOLD_START_COL + i, value=grade)
                else:
                    for i, lvl in enumerate(sorted_levels):
                        ws.cell(row=row_idx, column=THRESHOLD_START_COL + i, value=th.get(lvl['key'], 0))

                # Create Excel formula for score calculation
                score_formula = _create_score_formula(row_idx, kr['metric_type'], num_levels)
                ws.cell(row=row_idx, column=SCORE_COL, value=score_formula)

                # Create Excel formula for performance level
                perf_level_formula = _create_performance_level_formula(row_idx, num_levels)
                ws.cell(row=row_idx, column=LEVEL_COL, value=perf_level_formula)

                # Get cells for formatting
                score_cell = ws.cell(row=row_idx, column=SCORE_COL)
                level_cell = ws.cell(row=row_idx, column=LEVEL_COL)

                # Set number format for score (2 decimal places)
                score_cell.number_format = '0.00'

                # Apply BASE cell coloring based on initial calculated score
                # This ensures correct colors when file is first opened
                # Conditional formatting will override these when formulas recalculate
                base_color = colors.get(initial_level, 'd9534f')
                base_fill = PatternFill(start_color=base_color, end_color=base_color, fill_type='solid')

                score_cell.fill = base_fill
                score_cell.font = Font(bold=True, color='FFFFFF')
                score_cell.alignment = Alignment(horizontal='center', vertical='center')

                level_cell.fill = base_fill
                level_cell.font = Font(bold=True, color='FFFFFF')
                level_cell.alignment = Alignment(horizontal='center', vertical='center')

                # Apply weight column styling (objective weight and KR weight)
                obj_weight_cell = ws.cell(row=row_idx, column=3)
                obj_weight_cell.fill = PatternFill(start_color='fef3c7', end_color='fef3c7', fill_type='solid')
                obj_weight_cell.font = Font(bold=True, color='d97706')

                kr_weight_cell = ws.cell(row=row_idx, column=5)
                kr_weight_cell.fill = PatternFill(start_color='fef3c7', end_color='fef3c7', fill_type='solid')
                kr_weight_cell.font = Font(bold=True, color='d97706')

                row_idx += 1

            # Merge objective cells if there are multiple KRs
            if len(kr_list) > 1:
                obj_end_row = row_idx - 1
                # Merge objective name
                ws.merge_cells(start_row=obj_start_row, start_column=2, end_row=obj_end_row, end_column=2)
                # Merge objective weight
                ws.merge_cells(start_row=obj_start_row, start_column=3, end_row=obj_end_row, end_column=3)

            # Apply formatting to objective cell
            obj_cell = ws.cell(row=obj_start_row, column=2)
            obj_cell.alignment = Alignment(horizontal='center', vertical='center')
            obj_cell.font = Font(bold=True)

            # Add darker border after each objective (bottom of last row)
            obj_end_row = row_idx - 1
            thick_bottom = Side(style='medium', color='000000')
            for col in range(1, TOTAL_COLS + 1):
                cell = ws.cell(row=obj_end_row, column=col)
                cell.border = Border(
                    left=cell.border.left if cell.border else None,
                    right=cell.border.right if cell.border else None,
                    top=cell.border.top if cell.border else None,
                    bottom=thick_bottom
                )

        # Merge department cells across ALL objectives in this department
        dept_end_row = row_idx - 1
        if dept_end_row > dept_start_row:
            ws.merge_cells(start_row=dept_start_row, start_column=1, end_row=dept_end_row, end_column=1)

        # Apply formatting to department cell
        dept_cell = ws.cell(row=dept_start_row, column=1)
        dept_cell.alignment = Alignment(horizontal='center', vertical='center')
        dept_cell.font = Font(bold=True)

    # Apply conditional formatting to score and performance level columns (only if there's data)
    if row_idx > 2:
        _apply_conditional_formatting(ws, row_idx, colors, num_levels)

    # Set column widths
    ws.column_dimensions['A'].width = 20  # Department
    ws.column_dimensions['B'].width = 30  # Objective
    ws.column_dimensions['C'].width = 12  # Objective Weight
    ws.column_dimensions['D'].width = 35  # Key Result
    ws.column_dimensions['E'].width = 12  # KR Weight
    ws.column_dimensions['F'].width = 15  # Type
    ws.column_dimensions['G'].width = 10  # Actual
    ws.column_dimensions['H'].width = 8   # Unit

    # Dynamic threshold columns
    for i in range(num_levels):
        col_letter = chr(ord('I') + i)
        ws.column_dimensions[col_letter].width = 12

    # Score and Level columns
    score_col_letter = chr(ord('A') + SCORE_COL - 1)
    level_col_letter = chr(ord('A') + LEVEL_COL - 1)
    ws.column_dimensions[score_col_letter].width = 10
    ws.column_dimensions[level_col_letter].width = 18

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def import_from_excel(file_content):
    """Import OKR data from Excel file.

    Expected columns: Department, Objective, Objective Weight, Key Result, KR Weight,
    Type, Actual, Unit, [Threshold columns...], Score, Performance Level

    Returns: (departments_list, success_flag, message)
    """
    try:
        wb = load_workbook(filename=BytesIO(file_content), data_only=True)
        ws = wb.active

        # Get header row to determine column mapping
        headers = [cell.value for cell in ws[1]]
        if not headers or len(headers) < 8:
            return None, False, t('import_error')

        # Get current score levels config to determine threshold columns
        config = get_score_levels_config()
        sorted_levels = sorted(config['levels'], key=lambda x: x['order'])
        num_levels = len(sorted_levels)

        # Column indices (0-based)
        COL_DEPT = 0
        COL_OBJ = 1
        COL_OBJ_WEIGHT = 2
        COL_KR = 3
        COL_KR_WEIGHT = 4
        COL_TYPE = 5
        COL_ACTUAL = 6
        COL_UNIT = 7
        COL_THRESHOLD_START = 8

        departments = {}
        current_dept_name = None
        current_obj_name = None
        current_dept = None
        current_obj = None

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None or cell == '' for cell in row[:4]):
                continue  # Skip empty rows

            # Get values, handling merged cells (empty values inherit from above)
            dept_name = row[COL_DEPT] if row[COL_DEPT] else current_dept_name
            obj_name = row[COL_OBJ] if row[COL_OBJ] else current_obj_name
            obj_weight_raw = row[COL_OBJ_WEIGHT] if row[COL_OBJ_WEIGHT] else None
            kr_name = row[COL_KR]
            kr_weight_raw = row[COL_KR_WEIGHT] if len(row) > COL_KR_WEIGHT else None
            type_raw = row[COL_TYPE] if len(row) > COL_TYPE else None
            actual_raw = row[COL_ACTUAL] if len(row) > COL_ACTUAL else None
            unit = row[COL_UNIT] if len(row) > COL_UNIT else ''

            if not dept_name or not obj_name or not kr_name:
                continue  # Skip rows without essential data

            # Parse weights (remove % if present)
            def parse_weight(val):
                if val is None:
                    return 0
                if isinstance(val, str):
                    val = val.replace('%', '').strip()
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return 0

            obj_weight = parse_weight(obj_weight_raw)
            kr_weight = parse_weight(kr_weight_raw)

            # Determine metric type
            metric_type = 'higher_better'  # default
            if type_raw:
                type_str = str(type_raw).lower()
                if 'qualitative' in type_str or 'a-e' in type_str or 'a/b/c' in type_str:
                    metric_type = 'qualitative'
                elif 'lower' in type_str or '↓' in type_str:
                    metric_type = 'lower_better'
                elif 'higher' in type_str or '↑' in type_str:
                    metric_type = 'higher_better'

            # Parse actual value
            actual = 0
            if actual_raw is not None:
                if metric_type == 'qualitative':
                    actual = str(actual_raw).upper().strip() if actual_raw else 'E'
                    if actual not in ['A', 'B', 'C', 'D', 'E']:
                        actual = 'E'
                else:
                    try:
                        actual = float(actual_raw)
                    except (ValueError, TypeError):
                        actual = 0

            # Parse thresholds from threshold columns
            thresholds = {}
            for i, lvl in enumerate(sorted_levels):
                th_col = COL_THRESHOLD_START + i
                if len(row) > th_col and row[th_col] is not None:
                    try:
                        if metric_type == 'qualitative':
                            # For qualitative, thresholds are grade letters - skip
                            pass
                        else:
                            thresholds[lvl['key']] = float(row[th_col])
                    except (ValueError, TypeError):
                        pass

            # Create or get department
            if dept_name != current_dept_name:
                current_dept_name = dept_name
                if dept_name not in departments:
                    departments[dept_name] = {
                        'id': str(uuid.uuid4()),
                        'name': dept_name,
                        'objectives': []
                    }
                current_dept = departments[dept_name]
                current_obj = None
                current_obj_name = None

            # Create or get objective
            if obj_name != current_obj_name:
                current_obj_name = obj_name
                # Check if objective already exists in this department
                existing_obj = None
                for obj in current_dept['objectives']:
                    if obj['name'] == obj_name:
                        existing_obj = obj
                        break

                if existing_obj:
                    current_obj = existing_obj
                    # Update weight if provided
                    if obj_weight > 0:
                        current_obj['weight'] = obj_weight
                else:
                    current_obj = {
                        'id': str(uuid.uuid4()),
                        'name': obj_name,
                        'weight': obj_weight,
                        'key_results': []
                    }
                    current_dept['objectives'].append(current_obj)

            # Create key result
            kr = {
                'id': str(uuid.uuid4()),
                'name': kr_name,
                'metric_type': metric_type,
                'unit': str(unit) if unit else '',
                'description': '',
                'weight': kr_weight,
                'thresholds': thresholds,
                'actual': actual
            }
            current_obj['key_results'].append(kr)

        # Convert departments dict to list
        departments_list = list(departments.values())

        if not departments_list:
            return None, False, t('import_error')

        return departments_list, True, t('import_success')

    except Exception as e:
        return None, False, f"{t('import_error')}: {str(e)}"


def inject_global_css():
    """Inject custom CSS for professional enterprise appearance"""
    st.markdown("""
    <style>
    /* Import professional fonts */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

    /* Global font and background */
    html, body, [class*="css"] {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    }

    .main {
        background: linear-gradient(180deg, #f8fafc 0%, #f1f5f9 100%);
    }
    ß

    /* ===== AGGRESSIVE TOP SPACE REMOVAL ===== */
    /* Remove default Streamlit padding */
    .main .block-container {
        padding-top: 0 !important;
        padding-bottom: 0rem;
        padding-left: 1.5rem;
        padding-right: 1.5rem;
        max-width: 100%;
    }

    /* Target the root app container */
    .stApp {
        margin-top: -80px !important;
    }

    /* Alternative: use negative margin on main content */
    [data-testid="stAppViewContainer"] {
        margin-top: 0 !important;
        padding-top: 0 !important;
    }

    [data-testid="stAppViewContainer"] > .main {
        padding-top: 0 !important;
    }

    /* Hide Streamlit branding and COMPLETELY remove header space */
    #MainMenu {display: none !important;}
    footer {display: none !important;}
    header {display: none !important; height: 0 !important;}

    /* Remove header element completely */
    [data-testid="stHeader"] {
        display: none !important;
        height: 0 !important;
    }

    .stApp > header {
        display: none !important;
        height: 0 !important;
    }

    /* Remove any top decoration/toolbar */
    [data-testid="stToolbar"] {
        display: none !important;
    }

    /* Remove deploy button area */
    [data-testid="stDecoration"] {
        display: none !important;
    }

    /* Remove top bar/status bar */
    [data-testid="stStatusWidget"] {
        display: none !important;
    }

    /* Ensure no top margin on first element */
    .element-container:first-child {
        margin-top: 0 !important;
    }

    /* Target iframe container if embedded */
    .stApp iframe {
        margin-top: 0 !important;
    }
    /* ===== END TOP SPACE REMOVAL ===== */

    /* Professional buttons */
    .stButton>button {
        font-family: 'Inter', sans-serif;
        font-weight: 500;
        border-radius: 8px;
        border: 1px solid #e4e7ec;
        transition: all 0.2s ease;
        box-shadow: 0 1px 3px rgba(0,0,0,0.08);
    }

    .stButton>button:hover {
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(0,0,0,0.15);
    }

    .stButton>button[kind="primary"] {
        background: linear-gradient(135deg, #0066cc 0%, #0052a3 100%);
        border: none;
        color: white;
    }

    /* Download button styling */
    .stDownloadButton>button {
        font-family: 'Inter', sans-serif;
        font-weight: 500;
        border-radius: 8px;
        background: linear-gradient(135deg, #059669 0%, #047857 100%);
        border: none;
        color: white;
        transition: all 0.2s ease;
    }

    .stDownloadButton>button:hover {
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(5, 150, 105, 0.4);
    }

    /* Expander styling */
    .streamlit-expanderHeader {
        font-family: 'Inter', sans-serif;
        font-weight: 600;
        background: #f8fafc;
        border-radius: 8px;
        padding: 0.75rem 1rem;
        border: 1px solid #e4e7ec;
    }

    .streamlit-expanderContent {
        border: 1px solid #e4e7ec;
        border-top: none;
        border-radius: 0 0 8px 8px;
        background: white;
    }

    /* Input fields */
    .stTextInput>div>div>input, .stNumberInput>div>div>input, .stSelectbox>div>div {
        font-family: 'Inter', sans-serif;
        border-radius: 8px;
        border: 1px solid #e4e7ec;
    }

    .stTextInput>div>div>input:focus, .stNumberInput>div>div>input:focus {
        border-color: #0066cc;
        box-shadow: 0 0 0 3px rgba(0, 102, 204, 0.1);
    }

    /* Radio buttons */
    .stRadio>div {
        background: white;
        padding: 0.5rem;
        border-radius: 8px;
        border: 1px solid #e4e7ec;
    }

    /* Metrics and stats cards */
    [data-testid="stMetricValue"] {
        font-family: 'Inter', sans-serif;
        font-weight: 700;
    }

    /* Custom scrollbar */
    ::-webkit-scrollbar {
        width: 8px;
        height: 8px;
    }
    ::-webkit-scrollbar-track {
        background: #f1f5f9;
        border-radius: 4px;
    }
    ::-webkit-scrollbar-thumb {
        background: #cbd5e1;
        border-radius: 4px;
    }
    ::-webkit-scrollbar-thumb:hover {
        background: #94a3b8;
    }

    /* Divider styling */
    hr {
        border: none;
        height: 1px;
        background: linear-gradient(90deg, transparent, #e4e7ec, transparent);
        margin: 1rem 0;
    }

    /* ===== STICKY SIDEBAR ===== */
    /* Make the sidebar column sticky */
    [data-testid="stHorizontalBlock"] > div:first-child {
        position: sticky;
        top: 0;
        align-self: flex-start;
        max-height: 100vh;
        overflow-y: auto;
    }

    /* Ensure proper scrolling behavior */
    [data-testid="stHorizontalBlock"] {
        align-items: flex-start !important;
    }
    /* ===== END STICKY SIDEBAR ===== */

    /* Alert/warning styling */
    .stAlert {
        border-radius: 8px;
        border: none;
    }
    </style>
    """, unsafe_allow_html=True)


@st.dialog("Score Level Settings", width="large")
def score_levels_settings_dialog():
    """Modal dialog for configuring score levels."""
    # Initialize dialog-specific session state for editing levels
    if 'dialog_levels_config' not in st.session_state:
        st.session_state.dialog_levels_config = None

    # Load current config into dialog state if not already loaded
    if st.session_state.dialog_levels_config is None:
        import copy
        st.session_state.dialog_levels_config = copy.deepcopy(get_score_levels_config())

    settings_config = st.session_state.dialog_levels_config

    # Score Range Configuration
    st.markdown(f"### {t('score_range')}")
    range_col1, range_col2 = st.columns(2)
    with range_col1:
        new_min = st.number_input(t("min_score"), value=settings_config['min_score'], step=0.25, key="dlg_cfg_min_score")
    with range_col2:
        new_max = st.number_input(t("max_score"), value=settings_config['max_score'], step=0.25, key="dlg_cfg_max_score")

    st.markdown("---")
    st.markdown(f"### {t('performance_levels')}")

    # Display existing levels
    settings_sorted_levels = sorted(settings_config['levels'], key=lambda x: x['order'])

    levels_to_update = []
    levels_to_delete = []

    for idx, level in enumerate(settings_sorted_levels):
        with st.container(border=True):
            lc1, lc2, lc3, lc4 = st.columns([2.5, 1.5, 1, 0.5])
            with lc1:
                new_name = st.text_input(
                    f"{t('level_name')} (EN)",
                    value=level['names'].get('en', ''),
                    key=f"dlg_lvl_name_{level['key']}"
                )
            with lc2:
                new_threshold = st.number_input(
                    t("level_threshold"),
                    value=level['threshold'],
                    step=0.01,
                    format="%.2f",
                    key=f"dlg_lvl_th_{level['key']}"
                )
            with lc3:
                new_color = st.color_picker(
                    t("level_color"),
                    value=level['color'],
                    key=f"dlg_lvl_color_{level['key']}"
                )
            with lc4:
                st.write("")  # Spacer
                if len(settings_sorted_levels) > 2:  # Keep at least 2 levels
                    if st.button("🗑️", key=f"dlg_del_lvl_{level['key']}", help=t("delete_level")):
                        # Remove from dialog config directly
                        st.session_state.dialog_levels_config['levels'] = [
                            lvl for lvl in st.session_state.dialog_levels_config['levels']
                            if lvl['key'] != level['key']
                        ]
                        # Re-order remaining levels
                        for i, lvl in enumerate(sorted(st.session_state.dialog_levels_config['levels'], key=lambda x: x['threshold'])):
                            lvl['order'] = i
                        st.rerun()

            levels_to_update.append({
                "key": level['key'],
                "order": idx,
                "threshold": new_threshold,
                "color": new_color,
                "names": {
                    "en": new_name,
                    "ru": level['names'].get('ru', new_name),
                    "uz": level['names'].get('uz', new_name)
                }
            })

    # Add new level button
    if st.button(f"➕ {t('add_level')}", key="dlg_add_new_level"):
        # Generate new level
        existing_keys = [lvl['key'] for lvl in settings_config['levels']]
        new_key_num = 1
        while f"level_{new_key_num}" in existing_keys:
            new_key_num += 1
        new_key = f"level_{new_key_num}"

        new_level = {
            "key": new_key,
            "order": len(settings_config['levels']),
            "threshold": round((new_min + new_max) / 2, 2),
            "color": "#808080",
            "names": {"en": "New Level", "ru": "Новый уровень", "uz": "Yangi daraja"}
        }
        st.session_state.dialog_levels_config['levels'].append(new_level)
        st.rerun()

    st.markdown("---")
    st.markdown(f"### {t('grade_mapping')}")

    # Qualitative grade mapping - use current dialog levels
    level_options = [lvl['key'] for lvl in settings_sorted_levels]
    level_labels = {lvl['key']: lvl['names'].get('en', lvl['key']) for lvl in settings_sorted_levels}

    grade_cols = st.columns(5)
    new_mapping = {}
    for i, grade in enumerate(['A', 'B', 'C', 'D', 'E']):
        with grade_cols[i]:
            current = settings_config['qualitative_mapping'].get(grade, level_options[-1] if level_options else '')
            if current not in level_options and level_options:
                current = level_options[-1]
            selected = st.selectbox(
                f"{t('qualitative_grade')} {grade}",
                options=level_options,
                format_func=lambda x, labels=level_labels: labels.get(x, x),
                index=level_options.index(current) if current in level_options else 0,
                key=f"dlg_grade_map_{grade}"
            )
            new_mapping[grade] = selected

    st.markdown("---")

    # Save and Reset buttons
    save_col, reset_col, cancel_col = st.columns(3)
    with save_col:
        if st.button(f"💾 {t('save_settings')}", type="primary", key="dlg_save_level_settings", use_container_width=True):
            # Build new config from form values
            new_config = {
                "min_score": new_min,
                "max_score": new_max,
                "levels": levels_to_update,
                "qualitative_mapping": new_mapping
            }

            # Re-sort and re-order levels by threshold
            new_config['levels'] = sorted(new_config['levels'], key=lambda x: x['threshold'])
            for i, lvl in enumerate(new_config['levels']):
                lvl['order'] = i

            # Validate
            is_valid, error_msg = validate_score_levels_config(new_config)
            if is_valid:
                # Get valid level keys from new config
                valid_level_keys = {lvl['key'] for lvl in new_config['levels']}

                # Clean up KR thresholds - remove keys for deleted levels
                for dept in st.session_state.departments:
                    for obj in dept.get('objectives', []):
                        for kr in obj.get('key_results', []):
                            if 'thresholds' in kr:
                                # Remove threshold keys that no longer exist in config
                                kr['thresholds'] = {
                                    k: v for k, v in kr['thresholds'].items()
                                    if k in valid_level_keys
                                }

                st.session_state.score_levels_config = new_config
                st.session_state.dialog_levels_config = None  # Clear dialog state
                save_data()
                st.success(t("settings_saved"))
                st.rerun()
            else:
                st.error(f"{t('invalid_config')}: {error_msg}")

    with reset_col:
        if st.button(f"🔄 {t('reset_defaults')}", key="dlg_reset_level_settings", use_container_width=True):
            import copy
            st.session_state.dialog_levels_config = copy.deepcopy(DEFAULT_SCORE_LEVELS_CONFIG)
            st.rerun()

    with cancel_col:
        if st.button(f"❌ {t('cancel')}", key="dlg_cancel_settings", use_container_width=True):
            st.session_state.dialog_levels_config = None  # Clear dialog state
            st.rerun()


def main():
    st.set_page_config(page_title="OKR Tracker", page_icon="🎯", layout="wide")
    inject_global_css()

    # Initialize
    if 'initialized' not in st.session_state:
        loaded_departments, loaded_lang, loaded_config = load_data()
        st.session_state.departments = loaded_departments
        st.session_state.language = loaded_lang
        st.session_state.score_levels_config = loaded_config
        st.session_state.new_krs = []
        st.session_state.initialized = True

    # Language selector aligned to the right
    col_lang = st.columns([5, 1])[1]
    with col_lang:
        lang_options = {"en": "🇬🇧 EN", "ru": "🇷🇺 RU", "uz": "🇺🇿 UZ"}
        selected_lang = st.selectbox("Language", list(lang_options.keys()),
                                     format_func=lambda x: lang_options[x],
                                     index=list(lang_options.keys()).index(st.session_state.language),
                                     label_visibility="collapsed")
        if selected_lang != st.session_state.language:
            st.session_state.language = selected_lang
            save_data()
            st.rerun()

    # ===== SIDEBAR TOGGLE =====
    if 'sidebar_collapsed' not in st.session_state:
        st.session_state.sidebar_collapsed = False

    # Toggle button
    col_toggle, col_spacer = st.columns([0.1, 0.9])
    with col_toggle:
        if st.button("◀" if not st.session_state.sidebar_collapsed else "▶",
                     key="sidebar_toggle"):
            st.session_state.sidebar_collapsed = not st.session_state.sidebar_collapsed
            st.rerun()

    # ===== MAIN LAYOUT: SIDEBAR + DASHBOARD =====
    if st.session_state.sidebar_collapsed:
        # Sidebar is hidden, show only main content
        col_main = st.container()
    else:
        # Sidebar is visible, use two-column layout
        col_sidebar, col_main = st.columns([0.22, 0.78], gap="medium")

    if not st.session_state.sidebar_collapsed:
        with col_sidebar:
            # === SIDEBAR ===
            render_sidebar(st.session_state.departments)

            # Department navigation with dropdown/combo box
            st.markdown(
                f"<h3 style='font-size:14px; color:{THEME['text_secondary']}; text-transform:uppercase; letter-spacing:1px; margin:25px 0 12px 0;'>🏢 {t('departments')}</h3>",
                unsafe_allow_html=True)

            if 'selected_dept_filter' not in st.session_state:
                st.session_state.selected_dept_filter = t("all_departments")

            # Department selectbox/combo box
            if st.session_state.departments:
                dept_options = [t("all_departments")] + [d['name'] for d in st.session_state.departments]
                selected_dept_filter = st.selectbox(
                    t("select_department"),
                    dept_options,
                    index=dept_options.index(
                        st.session_state.selected_dept_filter) if st.session_state.selected_dept_filter in dept_options else 0,
                    label_visibility="collapsed",
                    key="dept_filter_select"
                )
                if selected_dept_filter != st.session_state.selected_dept_filter:
                    st.session_state.selected_dept_filter = selected_dept_filter
                    st.rerun()

            # View mode switcher
            st.markdown(
                f"<h3 style='font-size:14px; color:{THEME['text_secondary']}; text-transform:uppercase; letter-spacing:1px; margin:25px 0 12px 0;'> {t('view_mode')}</h3>",
                unsafe_allow_html=True)
            view_mode = st.radio(
                "View",
                ["Grid", "Full"],
                horizontal=False,
                label_visibility="collapsed",
                key="view_mode_radio"
            )
            if view_mode.lower() != st.session_state.get('view_mode', 'grid'):
                st.session_state.view_mode = view_mode.lower()

            # Action buttons
            st.markdown(
                f"<h3 style='font-size:14px; color:{THEME['text_secondary']}; text-transform:uppercase; letter-spacing:1px; margin:25px 0 12px 0;'>⚙️ {t('actions')}</h3>",
                unsafe_allow_html=True)

            if st.button("💾 " + t("save_data"), use_container_width=True, type="primary"):
                save_data()
                st.success(t("data_saved"))

            if st.button(t("load_data"), use_container_width=True):
                dept, lang, config = load_data()
                if dept:
                    st.session_state.departments = dept
                    st.session_state.language = lang
                    st.session_state.score_levels_config = config
                    st.success(t("data_loaded"))
                    st.rerun()
                else:
                    st.warning(t("no_data"))

            # Export button - Excel only
            excel_data = export_to_excel(st.session_state.departments)
            st.download_button(
                label="" + t("export_excel"),
                data=excel_data,
                file_name="okr_export.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

            # Import from Excel
            uploaded_file = st.file_uploader(
                t("import_excel"),
                type=['xlsx', 'xls'],
                key="excel_import_uploader"
            )
            if uploaded_file is not None:
                file_content = uploaded_file.read()
                departments, success, message = import_from_excel(file_content)
                if success:
                    st.session_state.departments = departments
                    save_data()
                    st.success(message)
                    st.rerun()
                else:
                    st.error(message)

            # ===== SCORE LEVEL SETTINGS =====
            st.markdown(
                f"<h3 style='font-size:14px; color:{THEME['text_secondary']}; text-transform:uppercase; letter-spacing:1px; margin:25px 0 12px 0;'>⚙️ {t('score_level_settings')}</h3>",
                unsafe_allow_html=True)

            if st.button(f"⚙️ {t('configure_score_levels')}", use_container_width=True):
                score_levels_settings_dialog()

    with col_main:
        # === MAIN DASHBOARD AREA ===

        # Performance scale legend - dynamic N columns based on config
        config = get_score_levels_config()
        sorted_levels = sorted(config['levels'], key=lambda x: x['order'])
        levels_dict = get_levels_dict()

        st.markdown(f"""
            <p style='font-size:12px; font-weight:600; margin:0 0 12px 0; color:{THEME['text_secondary']}; text-transform:uppercase; letter-spacing:1px;'> {t('performance_scale')}</p>
        """, unsafe_allow_html=True)

        cols = st.columns(len(sorted_levels))
        for i, level_cfg in enumerate(sorted_levels):
            level = levels_dict[level_cfg['key']]
            with cols[i]:
                pct_range = f"{score_to_percentage(level['min'])}%-{score_to_percentage(level['max'])}%"
                st.markdown(f"""
                    <div style='background:linear-gradient(135deg, {level['color']} 0%, {level['color']}dd 100%); color:white; padding:12px 10px; border-radius:10px; text-align:center; box-shadow:0 2px 8px {level['color']}30; margin-bottom:16px;'>
                        <div style='font-size:12px; font-weight:700; margin-bottom:4px;'>{get_level_label(level_cfg['key'])}</div>
                        <div style='font-size:11px; opacity:0.9;'>{level['min']:.2f} - {level['max']:.2f}</div>
                        <div style='font-size:10px; opacity:0.75; margin-top:2px;'>{pct_range}</div>
                    </div>
                """, unsafe_allow_html=True)

        # ===== CREATE DEPARTMENT =====
        with st.expander(t("create_department"), expanded=len(st.session_state.departments) == 0):
            new_dept_name = st.text_input(t("department_name"), key="new_dept_name")
            if st.button(t("create_department") + " ✅", key="create_dept_btn"):
                if new_dept_name.strip():
                    st.session_state.departments.append({
                        "id": str(uuid.uuid4()),
                        "name": new_dept_name.strip(),
                        "objectives": []
                    })
                    save_data()
                    st.rerun()
                else:
                    st.error(t("enter_dept_name"))

        # ===== CREATE OBJECTIVE =====
        with st.expander(t("create_objective"), expanded=False):
            if not st.session_state.departments:
                st.warning(t("no_departments"))
            else:
                # Department selector
                dept_options = {dept['id']: dept['name'] for dept in st.session_state.departments}
                selected_dept_id = st.selectbox(
                    t("select_department"),
                    options=list(dept_options.keys()),
                    format_func=lambda x: dept_options[x],
                    key="selected_dept_for_obj"
                )

                # Objective name and weight
                obj_col1, obj_col2 = st.columns([4, 1])
                with obj_col1:
                    new_obj_name = st.text_input(t("objective_name"), key="new_obj_name")
                with obj_col2:
                    new_obj_weight = st.number_input(t("objective_weight"), min_value=0, max_value=100, value=0,
                                                     key="new_obj_weight",
                                                     help="Weight of this objective within department (0-100%)")

                st.markdown(f"#### {t('add_key_results')}")

                # KR basic info (without weight)
                c1, c2, c3 = st.columns([3, 1.5, 1])
                with c1:
                    kr_name = st.text_input(t("kr_name"), key="kr_name_input")
                with c2:
                    kr_type = st.selectbox(t("type"), ["higher_better", "lower_better", "qualitative"],
                                           format_func=lambda x: t(x),
                                           key="kr_type_input")
                with c3:
                    kr_unit = st.text_input(t("unit"), value="%" if kr_type != "qualitative" else "",
                                            key="kr_unit_input",
                                            disabled=(kr_type == "qualitative"))

                # Description field for tooltip
                kr_description = st.text_area(t("kr_description"), placeholder=t("kr_description_placeholder"),
                                              key="kr_description_input", height=68)

                # Weight field
                kr_weight = st.number_input(
                    t("kr_weight"),
                    value=0.0,
                    min_value=0.0,
                    max_value=100.0,
                    step=1.0,
                    key="kr_weight_input"
                )

                # Show thresholds only for quantitative metrics - dynamic levels
                level_config = get_score_levels_config()
                sorted_lvls = sorted(level_config['levels'], key=lambda x: x['order'])

                if kr_type != "qualitative":
                    st.markdown(f"**{t('thresholds')}:**")
                    threshold_cols = st.columns(len(sorted_lvls))
                    threshold_values = {}

                    # Calculate default values (evenly distributed 0-100)
                    for idx, lvl in enumerate(sorted_lvls):
                        default_val = (100.0 / (len(sorted_lvls) - 1)) * idx if len(sorted_lvls) > 1 else 0.0
                        with threshold_cols[idx]:
                            st.markdown(f"<small style='color:{lvl['color']};'>● {lvl['threshold']:.2f}</small>", unsafe_allow_html=True)
                            threshold_values[lvl['key']] = st.number_input(
                                get_level_label(lvl['key']),
                                value=round(default_val, 1),
                                key=f"th_{lvl['key']}"
                            )
                else:
                    # Qualitative KR info box - dynamic text
                    qual_mapping = level_config.get('qualitative_mapping', {})
                    grade_text_parts = []
                    for grade in ['A', 'B', 'C', 'D', 'E']:
                        if grade in qual_mapping:
                            level_key = qual_mapping[grade]
                            for lvl in level_config['levels']:
                                if lvl['key'] == level_key:
                                    grade_text_parts.append(f"{grade}={lvl['threshold']:.2f} ({get_level_label(level_key)})")
                                    break
                    st.info(f" Qualitative KRs use A/B/C/D/E grades: {', '.join(grade_text_parts)}")
                    threshold_values = {lvl['key']: 0.0 for lvl in sorted_lvls}

                if st.button(t("add_kr")):
                    if kr_name.strip():
                        st.session_state.new_krs.append({
                            "id": str(uuid.uuid4()), "name": kr_name.strip(), "metric_type": kr_type,
                            "unit": "" if kr_type == "qualitative" else kr_unit,
                            "description": kr_description.strip(),
                            "weight": kr_weight,
                            "thresholds": threshold_values,
                            "actual": "E" if kr_type == "qualitative" else 0.0
                        })
                        st.rerun()

                if st.session_state.new_krs:
                    st.markdown(f"**{t('added_krs')}:**")

                    for i, kr in enumerate(st.session_state.new_krs):
                        col1, col2, col3 = st.columns([4, 1, 1])
                        with col1:
                            if kr['metric_type'] == "qualitative":
                                icon = "⭐"
                            else:
                                icon = "↑" if kr['metric_type'] == "higher_better" else "↓"
                            kr_w = kr.get('weight', 0)
                            st.write(f"**KR{i + 1}: {kr['name']}** ({icon}) - {t('weight')}: {kr_w}%")
                        with col2:
                            if st.button(f"❌", key=f"rm_{kr['id']}"):
                                st.session_state.new_krs = [k for k in st.session_state.new_krs if k['id'] != kr['id']]
                                st.rerun()

                if st.button(t("create"), type="primary"):
                    if new_obj_name.strip() and st.session_state.new_krs:
                        # Find the department and add the objective
                        for dept in st.session_state.departments:
                            if dept['id'] == selected_dept_id:
                                dept['objectives'].append({
                                    "id": str(uuid.uuid4()), "name": new_obj_name.strip(),
                                    "weight": new_obj_weight,
                                    "key_results": st.session_state.new_krs.copy()
                                })
                                break
                        st.session_state.new_krs = []
                        save_data()
                        st.rerun()
                    else:
                        st.error(t("enter_name_error"))

        # Display objectives
        if st.session_state.departments:
            view_mode = st.session_state.get('view_mode', 'grid')
            dept_filter = st.session_state.get('selected_dept_filter', t("all_departments"))

            # Filter departments
            if dept_filter == t("all_departments"):
                departments_to_show = st.session_state.departments
            else:
                departments_to_show = [d for d in st.session_state.departments if d['name'] == dept_filter]

            for dept_idx, department in enumerate(st.session_state.departments):
                if dept_filter != t("all_departments") and department['name'] != dept_filter:
                    continue

                # Get actual index in full list
                actual_dept_idx = st.session_state.departments.index(department)

                # Department header
                st.markdown(
                    f"<div style='margin:20px 0 15px 0; padding-bottom:8px; border-bottom:2px solid {THEME['card_border']};'><h2 style='margin:0; font-size:20px; color:{THEME['text_primary']}; font-weight:600;'>📁 {department['name']}</h2></div>",
                    unsafe_allow_html=True)

                objectives = department.get('objectives', [])

                if not objectives:
                    st.info(t("no_objectives_yet"))
                    if st.button(t("delete_department"), key=f"del_dept_{department['id']}", type="secondary"):
                        st.session_state.departments = [d for d in st.session_state.departments if
                                                        d['id'] != department['id']]
                        save_data()
                        st.rerun()
                    continue

                # Render objectives as cards
                if view_mode == 'grid':
                    # Grid layout: 2 columns
                    for row_start in range(0, len(objectives), 2):
                        cols = st.columns(2, gap="medium")
                        for col_idx in range(2):
                            obj_idx = row_start + col_idx
                            if obj_idx < len(objectives):
                                with cols[col_idx]:
                                    render_objective_card(
                                        objectives[obj_idx],
                                        actual_dept_idx,
                                        obj_idx,
                                        compact=True
                                    )
                else:
                    # Full view: single column
                    for obj_idx, objective in enumerate(objectives):
                        render_objective_card(
                            objective,
                            actual_dept_idx,
                            obj_idx,
                            compact=False
                        )

                # Delete department button
                st.markdown("---")
                if st.button(t("delete_department") + f" '{department['name']}'",
                             key=f"del_dept_end_{department['id']}", type="secondary"):
                    st.session_state.departments = [d for d in st.session_state.departments if
                                                    d['id'] != department['id']]
                    save_data()
                    st.rerun()

        st.markdown("</div>", unsafe_allow_html=True)

    #  Empty state handled in main area
    if not st.session_state.departments:
        with col_main:
            st.info(t("create_first"))
            if st.button(t("load_demo")):
                # Create departments with demo objectives
                st.session_state.departments = [{
                    "id": str(uuid.uuid4()),
                    "name": "PMO - Project Management Office",
                    "objectives": [
                        # Цель 1: Обеспечить своевременную реализацию проектов (20%)
                        {
                            "id": str(uuid.uuid4()),
                            "name": "Цель 1: Обеспечить своевременную реализацию проектов",
                            "weight": 20,
                            "key_results": [
                                {"id": str(uuid.uuid4()),
                                 "name": "KR1.1 Проекты завершенные в срок (% от кол-ва проектов)",
                                 "metric_type": "higher_better", "unit": "%", "weight": 40,
                                 "description": "Процент проектов, которые были завершены в установленные сроки. Измеряется как отношение количества проектов, завершенных вовремя, к общему количеству проектов.",
                                 "thresholds": {"below": 50, "meets": 60, "good": 80, "exceptional": 120}, "actual": 0},
                                {"id": str(uuid.uuid4()), "name": "KR1.2 Задачи в JIRA, завершенные в срок (%)",
                                 "metric_type": "higher_better", "unit": "%", "weight": 35,
                                 "description": "Процент задач в системе JIRA, выполненных в установленные сроки без переносов дедлайнов.",
                                 "thresholds": {"below": 50, "meets": 65, "good": 95, "exceptional": 200}, "actual": 0},
                                {"id": str(uuid.uuid4()),
                                 "name": "KR1.3 Переносы сроков заверш задач в JIRA (% от общего кол-ва)",
                                 "metric_type": "lower_better", "unit": "%", "weight": 25,
                                 "description": "Процент задач, у которых были перенесены сроки выполнения. Чем меньше значение, тем лучше.",
                                 "thresholds": {"below": 30, "meets": 20, "good": 15, "exceptional": 0},
                                 "actual": 0},
                            ]
                        },
                        # Цель 2: Управление рисками и бюджетом проектов (20%)
                        {
                            "id": str(uuid.uuid4()),
                            "name": "Цель 2: Управление рисками и бюджетом проектов",
                            "weight": 20,
                            "key_results": [
                                {"id": str(uuid.uuid4()), "name": "KR2.1 Проекты в рамках бюджетов (% без превышения)",
                                 "metric_type": "higher_better", "unit": "%", "weight": 30,
                                 "thresholds": {"below": 50, "meets": 60, "good": 75, "exceptional": 100}, "actual": 0},
                                {"id": str(uuid.uuid4()),
                                 "name": "KR2.2 Неучтенные риски возникшие после начала проекта (кол-во)",
                                 "metric_type": "lower_better", "unit": "", "weight": 25,
                                 "thresholds": {"below": 10, "meets": 5, "good": 2, "exceptional": 0},
                                 "actual": 0},
                                {"id": str(uuid.uuid4()), "name": "KR2.3 Повысить точность оценки трудозатрат до 75%",
                                 "metric_type": "higher_better", "unit": "%", "weight": 25,
                                 "thresholds": {"below": 50, "meets": 75, "good": 80, "exceptional": 100}, "actual": 0},
                                {"id": str(uuid.uuid4()), "name": "KR2.4 Процент рисков с планами митигации (%)",
                                 "metric_type": "higher_better", "unit": "%", "weight": 20,
                                 "thresholds": {"below": 20, "meets": 50, "good": 60, "exceptional": 100}, "actual": 0},
                            ]
                        },
                        # Цель 3: Управление качеством и отчетность (20%)
                        {
                            "id": str(uuid.uuid4()),
                            "name": "Цель 3: Управление качеством и отчетность",
                            "weight": 20,
                            "key_results": [
                                {"id": str(uuid.uuid4()),
                                 "name": "KR3.1 Своевременность отчетов W,Q,Y, другие (задержка, дней)",
                                 "metric_type": "lower_better", "unit": " дней", "weight": 25,
                                 "thresholds": {"below": 5, "meets": 3, "good": 2, "exceptional": 0},
                                 "actual": 0},
                                {"id": str(uuid.uuid4()),
                                 "name": "KR3.2 Уровень использования ресурсов (resource utilization) %",
                                 "metric_type": "higher_better", "unit": "%", "weight": 25,
                                 "thresholds": {"below": 75, "meets": 85, "good": 90, "exceptional": 100}, "actual": 0},
                                {"id": str(uuid.uuid4()),
                                 "name": "KR3.3 Реагирование на изменения (Response time to changes) часы",
                                 "metric_type": "lower_better", "unit": " часов", "weight": 25,
                                 "thresholds": {"below": 5, "meets": 3, "good": 2, "exceptional": 0},
                                 "actual": 0},
                                {"id": str(uuid.uuid4()),
                                 "name": "KR3.4 Среднее время от инициации до завершения проекта (нед)",
                                 "metric_type": "lower_better", "unit": " нед", "weight": 25,
                                 "thresholds": {"below": 10, "meets": 8, "good": 6, "exceptional": 4},
                                 "actual": 0},
                            ]
                        },
                        # Цель 4: Усиление состава и человеческий капитал (10%) - includes qualitative KR
                        {
                            "id": str(uuid.uuid4()),
                            "name": "Цель 4: Усиление состава и человеческий капитал",
                            "weight": 10,
                            "key_results": [
                                {"id": str(uuid.uuid4()),
                                 "name": "KR4.1 Комплектация штата (6 свободных вакансий в штате)",
                                 "metric_type": "higher_better", "unit": "", "weight": 35,
                                 "thresholds": {"below": 2, "meets": 3, "good": 4, "exceptional": 6},
                                 "actual": 0},
                                {"id": str(uuid.uuid4()), "name": "KR4.2 Набор и подготовка стажеров (16 вакансий)",
                                 "metric_type": "higher_better", "unit": "", "weight": 35,
                                 "thresholds": {"below": 3, "meets": 6, "good": 10, "exceptional": 16},
                                 "actual": 0},
                                {"id": str(uuid.uuid4()), "name": "KR4.3 Качество развития сотрудников (оценка)",
                                 "metric_type": "qualitative", "unit": "", "weight": 30,
                                 "description": "Качественная оценка программы развития сотрудников. A=Отлично, B=Очень хорошо, C=Хорошо, D=Удовлетворительно, E=Неудовлетворительно",
                                 "thresholds": {"below": 0, "meets": 0, "good": 0, "exceptional": 0},
                                 "actual": "C"},
                            ]
                        },
                        # Цель 5: Улучшение продуктов (10%)
                        {
                            "id": str(uuid.uuid4()),
                            "name": "Цель 5: Улучшение продуктов",
                            "weight": 10,
                            "key_results": [
                                {"id": str(uuid.uuid4()),
                                 "name": "KR5.1 Увеличить долю проектов, связанных со стратегическими целями Банка, до 85%",
                                 "metric_type": "higher_better", "unit": "%", "weight": 30,
                                 "thresholds": {"below": 75, "meets": 85, "good": 90, "exceptional": 100}, "actual": 0},
                                {"id": str(uuid.uuid4()),
                                 "name": "KR5.2 % продуктов с повторными багами (Defect/error rate)",
                                 "metric_type": "lower_better", "unit": "%", "weight": 30,
                                 "thresholds": {"below": 20, "meets": 15, "good": 10, "exceptional": 0},
                                 "actual": 0},
                                {"id": str(uuid.uuid4()),
                                 "name": "KR5.3 Обеспечить участие 100% членов команды в обучении по Agile/Scrum",
                                 "metric_type": "higher_better", "unit": "%", "weight": 20,
                                 "thresholds": {"below": 80, "meets": 90, "good": 95, "exceptional": 100}, "actual": 0},
                                {"id": str(uuid.uuid4()),
                                 "name": "KR5.4 Провести 6 внутренних воркшопов по методологиям и новым технологиям",
                                 "metric_type": "higher_better", "unit": "", "weight": 20,
                                 "thresholds": {"below": 4, "meets": 6, "good": 7, "exceptional": 9},
                                 "actual": 0},
                            ]
                        },
                        # Цель 6: Системная и бизнес аналитика и ее автоматизация (20%)
                        {
                            "id": str(uuid.uuid4()),
                            "name": "Цель 6: Системная и бизнес аналитика и ее автоматизация",
                            "weight": 20,
                            "key_results": [
                                {"id": str(uuid.uuid4()),
                                 "name": "KR6.1 Уровень автоматизации процессов проектного управления",
                                 "metric_type": "higher_better", "unit": "%", "weight": 40,
                                 "thresholds": {"below": 75, "meets": 85, "good": 90, "exceptional": 100}, "actual": 0},
                                {"id": str(uuid.uuid4()),
                                 "name": "KR6.2 Качество описание бизнес процессов (изменение BPMN) %",
                                 "metric_type": "lower_better", "unit": "%", "weight": 30,
                                 "thresholds": {"below": 20, "meets": 15, "good": 10, "exceptional": 0},
                                 "actual": 0},
                                {"id": str(uuid.uuid4()),
                                 "name": "KR6.3 Процент изменений плана проекта после планирования",
                                 "metric_type": "lower_better", "unit": "%", "weight": 30,
                                 "thresholds": {"below": 20, "meets": 15, "good": 10, "exceptional": 0},
                                 "actual": 0},
                            ]
                        },
                    ]
                }]
                save_data()
                st.rerun()


if __name__ == "__main__":
    main()