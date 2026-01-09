import streamlit as st
import streamlit.components.v1 as components
import plotly.graph_objects as go
import pandas as pd
import json
import uuid
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

TRANSLATIONS = {
    "en": {
        "title": "OKR Performance Tracker",
        "performance_scale": "Performance Scale",
        "department": "Department",
        "department_name": "Department Name",
        "create_department": "➕ Create New Department",
        "select_department": "Select Department",
        "delete_department": "🗑️ Delete Department",
        "no_departments": "No departments. Create one first!",
        "create_objective": "➕ Create New Objective",
        "objective_name": "Objective Name",
        "objective": "Objective",
        "add_key_results": "Add Key Results",
        "kr_name": "KR Name",
        "type": "Type",
        "higher_better": "↑ Higher is better",
        "lower_better": "↓ Lower is better",
        "unit": "Unit",
        "thresholds": "Thresholds",
        "add_kr": "➕ Add KR",
        "added_krs": "Added Key Results",
        "remove": "Remove",
        "create": "✅ Create Objective",
        "enter_name_error": "Enter objective name and add at least one KR",
        "score": "🎯 Score",
        "add_kr_to_obj": "➕ Add KR to this Objective",
        "add": "➕ Add",
        "delete_objective": "🗑️ Delete Objective",
        "export_json": "Export JSON",
        "export_excel": "Export Excel",
        "save_data": "Save Data",
        "load_data": "📂 Load Data",
        "data_saved": "✅ Data saved!",
        "data_loaded": "✅ Data loaded!",
        "no_data": "No saved data found",
        "load_demo": "📋 Load Demo",
        "create_first": "👆 Create your first objective!",
        "language": "Language",
        "fact": "Fact",
        "actual": "Actual",
        "result": "Result",
        "delete": "Delete",
        "key_result": "Key Result",
        "no_krs": "No Key Results. Add some below.",
        "delete_krs": "🗑️ Delete Key Results",
        "performance_level": "Performance Level",
        "below": "Below", "meets": "Meets", "good": "Good", "very_good": "Very Good", "exceptional": "Exceptional",
        "view_grid": "Grid",
        "view_full": "Full",
    },
    "ru": {
        "title": "OKR Трекер",
        "performance_scale": "Шкала Оценки",
        "department": "Департамент",
        "department_name": "Название Департамента",
        "create_department": "➕ Создать Департамент",
        "select_department": "Выберите Департамент",
        "delete_department": "🗑️ Удалить Департамент",
        "no_departments": "Нет департаментов. Создайте сначала!",
        "create_objective": "➕ Создать Цель",
        "objective_name": "Название Цели",
        "objective": "Цель",
        "add_key_results": "Добавить Ключевые Результаты",
        "kr_name": "Название KR",
        "type": "Тип",
        "higher_better": "↑ Больше лучше",
        "lower_better": "↓ Меньше лучше",
        "unit": "Единица",
        "thresholds": "Пороги",
        "add_kr": "➕ Добавить KR",
        "added_krs": "Добавленные KR",
        "remove": "Удалить",
        "create": "✅ Создать",
        "enter_name_error": "Введите название и добавьте KR",
        "score": "🎯 Оценка",
        "add_kr_to_obj": "➕ Добавить KR",
        "add": "➕ Добавить",
        "delete_objective": "🗑️ Удалить Цель",
        "export_json": "Экспорт",
        "export_excel": "Экспорт Excel",
        "save_data": "Сохранить",
        "load_data": "📂 Загрузить",
        "data_saved": "✅ Сохранено!",
        "data_loaded": "✅ Загружено!",
        "no_data": "Нет данных",
        "load_demo": "📋 Демо",
        "create_first": "👆 Создайте цель!",
        "language": "Язык",
        "fact": "Факт",
        "actual": "Фактический",
        "result": "Результат",
        "delete": "Удалить",
        "key_result": "Ключевой Результат",
        "no_krs": "Нет KR.",
        "delete_krs": "🗑️ Удалить Ключевые Результаты",
        "performance_level": "Уровень Производительности",
        "below": "Ниже", "meets": "Соответствует", "good": "Хорошо", "very_good": "Очень хорошо", "exceptional": "Исключительно",
        "view_grid": "Сетка",
        "view_full": "Полный",
    },
    "uz": {
        "title": "OKR Трекер",
        "performance_scale": "Баҳолаш Шкаласи",
        "department": "Департамент",
        "department_name": "Департамент Номи",
        "create_department": "➕ Департамент Яратиш",
        "select_department": "Департамент Танланг",
        "delete_department": "🗑️ Департамент Ўчириш",
        "no_departments": "Департамент йўқ. Аввал яратинг!",
        "create_objective": "➕ Мақсад Яратиш",
        "objective_name": "Мақсад Номи",
        "objective": "Мақсад",
        "add_key_results": "Калит Натижалар Қўшиш",
        "kr_name": "KR Номи",
        "type": "Тури",
        "higher_better": "↑",
        "lower_better": "↓",
        "unit": "Бирлик",
        "thresholds": "Чегаралар",
        "add_kr": "➕ KR Қўшиш",
        "added_krs": "Қўшилган KR",
        "remove": "Ўчириш",
        "create": "✅ Яратиш",
        "enter_name_error": "Ном ва KR киритинг",
        "score": "🎯 Баҳо",
        "add_kr_to_obj": "➕ KR Қўшиш",
        "add": "➕ Қўшиш",
        "delete_objective": "🗑️ Ўчириш",
        "export_json": "Экспорт",
        "export_excel": "Excel Экспорт",
        "save_data": "Сақлаш",
        "load_data": "📂 Юклаш",
        "data_saved": "✅ Сақланди!",
        "data_loaded": "✅ Юкланди!",
        "no_data": "Маълумот йўқ",
        "load_demo": "📋 Демо",
        "create_first": "👆 Мақсад яратинг!",
        "language": "Тил",
        "fact": "Факт",
        "actual": "Ҳақиқий",
        "result": "Натижа",
        "delete": "Ўчириш",
        "key_result": "Калит Натижа",
        "no_krs": "KR йўқ.",
        "delete_krs": "🗑️ Калит Натижаларни Ўчириш",
        "performance_level": "Самарадорлик Даражаси",
        "below": "Ёмон", "meets": "Кутилган", "good": "Яхши", "very_good": "Жуда яхши", "exceptional": "Фантастик",
        "view_grid": "Тўр",
        "view_full": "Тўлиқ",
    }
}

LEVELS = {
    "below": {"min": 3.00, "max": 4.24, "color": "#d9534f"},
    "meets": {"min": 4.25, "max": 4.49, "color": "#f0ad4e"},
    "good": {"min": 4.50, "max": 4.74, "color": "#5cb85c"},
    "very_good": {"min": 4.75, "max": 4.99, "color": "#28a745"},
    "exceptional": {"min": 5.00, "max": 5.00, "color": "#1e7b34"},
}

DATA_FILE = "okr_data.json"


def t(key: str) -> str:
    lang = st.session_state.get('language', 'en')
    return TRANSLATIONS.get(lang, TRANSLATIONS['en']).get(key, key)


def get_level_label(level_key: str) -> str:
    return t(level_key)


def calculate_score(actual: float, metric_type: str, thresholds: dict) -> dict:
    below_th = thresholds['below']
    meets_th = thresholds['meets']
    good_th = thresholds['good']
    very_good_th = thresholds['very_good']
    exceptional_th = thresholds['exceptional']

    if metric_type == "higher_better":
        if actual >= exceptional_th:
            score = 5.00
            level = "exceptional"
        elif actual >= very_good_th:
            ratio = (actual - very_good_th) / max((exceptional_th - very_good_th), 1)
            score = 4.75 + ratio * 0.24
            level = "very_good"
        elif actual >= good_th:
            ratio = (actual - good_th) / max((very_good_th - good_th), 1)
            score = 4.50 + ratio * 0.24
            level = "good"
        elif actual >= meets_th:
            ratio = (actual - meets_th) / max((good_th - meets_th), 1)
            score = 4.25 + ratio * 0.24
            level = "meets"
        elif actual >= below_th:
            ratio = (actual - below_th) / max((meets_th - below_th), 1)
            score = 3.00 + ratio * 1.24
            level = "below"
        else:
            score = 3.00
            level = "below"
    else:
        # Lower is better
        if actual <= exceptional_th:
            score = 5.00
            level = "exceptional"
        elif actual <= very_good_th:
            ratio = 1 - (actual - exceptional_th) / max((very_good_th - exceptional_th), 1)
            score = 4.75 + ratio * 0.24
            level = "very_good"
        elif actual <= good_th:
            ratio = 1 - (actual - very_good_th) / max((good_th - very_good_th), 1)
            score = 4.50 + ratio * 0.24
            level = "good"
        elif actual <= meets_th:
            ratio = 1 - (actual - good_th) / max((meets_th - good_th), 1)
            score = 4.25 + ratio * 0.24
            level = "meets"
        elif actual <= below_th:
            ratio = 1 - (actual - meets_th) / max((below_th - meets_th), 1)
            score = 3.00 + ratio * 1.24
            level = "below"
        else:
            score = 3.00
            level = "below"

    return {"score": round(min(max(score, 3.0), 5.0), 2), "level": level, "level_info": LEVELS[level]}


def get_level_for_score(score: float) -> dict:
    if score >= 5.00:
        return {**LEVELS['exceptional'], "key": "exceptional"}
    elif score >= 4.75:
        return {**LEVELS['very_good'], "key": "very_good"}
    elif score >= 4.50:
        return {**LEVELS['good'], "key": "good"}
    elif score >= 4.25:
        return {**LEVELS['meets'], "key": "meets"}
    else:
        return {**LEVELS['below'], "key": "below"}


def score_to_percentage(score: float) -> float:
    return round(((score - 3.0) / 2.0) * 100, 1)


def create_gauge(score: float, compact: bool = False) -> str:
    """Returns HTML string with ECharts gauge"""
    import random
    percentage = score_to_percentage(score)
    level_info = get_level_for_score(score)
    level_label = get_level_label(level_info['key'])

    # Use unique ID to avoid conflicts when multiple gauges on page
    gauge_id = f"gauge_{random.randint(10000, 99999)}"

    # Compact mode settings
    height = 160 if compact else 280
    font_size = 14 if compact else 20
    label_size = 9 if compact else 11
    pointer_width = 8 if compact else 12
    axis_width = 18 if compact else 28

    html = f'''
    <div id="{gauge_id}" style="width: 100%; height: {height}px;"></div>
    <script src="https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js"></script>
    <script>
        var chart = echarts.init(document.getElementById('{gauge_id}'));
        var option = {{
            series: [{{
                type: 'gauge',
                min: 3,
                max: 5,
                splitNumber: 4,
                radius: '90%',
                center: ['50%', '60%'],
                startAngle: 180,
                endAngle: 0,
                axisLine: {{
                    lineStyle: {{
                        width: {axis_width},
                        color: [
                            [0.625, '#d9534f'],
                            [0.745, '#f0ad4e'],
                            [0.87, '#5cb85c'],
                            [0.995, '#28a745'],
                            [1, '#1e7b34']
                        ]
                    }}
                }},
                pointer: {{
                    icon: 'path://M12.8,0.7l12,40.1H0.7L12.8,0.7z',
                    length: '60%',
                    width: {pointer_width},
                    offsetCenter: [0, '-10%'],
                    itemStyle: {{
                        color: '#1a1a2e'
                    }}
                }},
                axisTick: {{
                    length: 5,
                    lineStyle: {{
                        color: 'auto',
                        width: 1
                    }}
                }},
                splitLine: {{
                    length: 10,
                    lineStyle: {{
                        color: 'auto',
                        width: 2
                    }}
                }},
                axisLabel: {{
                    color: '#444',
                    fontSize: {label_size},
                    distance: -35,
                    formatter: function (value) {{
                        return value.toFixed(1);
                    }}
                }},
                title: {{
                    show: false
                }},
                detail: {{
                    fontSize: {font_size},
                    offsetCenter: [0, '55%'],
                    valueAnimation: true,
                    formatter: function (value) {{
                        return value.toFixed(2) + '\\n({percentage}%)';
                    }},
                    color: '#2c3e50',
                    fontFamily: 'Arial',
                    fontWeight: 'bold'
                }},
                data: [{{
                    value: {score},
                    name: '{level_label}'
                }}]
            }}]
        }};
        chart.setOption(option);
    </script>
    '''
    return html


def save_data():
    data = {
        "departments": st.session_state.departments,
        "language": st.session_state.get('language', 'en')
    }
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)

            # Migration: convert old flat structure to department structure
            if 'objectives' in data and 'departments' not in data:
                # Old format: migrate to new format
                old_objectives = data.get('objectives', [])
                if old_objectives:
                    # Create a default department with all existing objectives
                    departments = [{
                        "id": str(uuid.uuid4()),
                        "name": "Default Department",
                        "objectives": old_objectives
                    }]
                else:
                    departments = []
                return departments, data.get('language', 'en')
            else:
                # New format
                return data.get('departments', []), data.get('language', 'en')
    return [], 'en'


def export_to_excel(departments):
    """Export OKR data to Excel with color-coded formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "OKR Export"

    # Define colors for performance levels
    colors = {
        'below': 'd9534f',
        'meets': 'f0ad4e',
        'good': '5cb85c',
        'very_good': '28a745',
        'exceptional': '1e7b34'
    }

    # Define header style
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Add headers
    headers = [t('department'), t('objective'), t('key_result'), t('actual'), t('unit'), t('below'), t('meets'), t('good'),
               t('very_good'), t('exceptional'), t('score').replace('🎯 ', ''), t('performance_level')]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Add data
    row_idx = 2
    for dept in departments:
        dept_name = dept['name']
        for obj in dept.get('objectives', []):
            obj_name = obj['name']
            start_row = row_idx  # Track starting row for this objective
            kr_list = obj.get('key_results', [])

            for kr in kr_list:
                # Calculate score
                result = calculate_score(kr['actual'], kr['metric_type'], kr['thresholds'])

                # Write data (department and objective names only in first row, will be merged later)
                ws.cell(row=row_idx, column=1, value=dept_name if row_idx == start_row else '')
                ws.cell(row=row_idx, column=2, value=obj_name if row_idx == start_row else '')
                ws.cell(row=row_idx, column=3, value=kr['name'])
                ws.cell(row=row_idx, column=4, value=kr['actual'])
                ws.cell(row=row_idx, column=5, value=kr['unit'])
                ws.cell(row=row_idx, column=6, value=kr['thresholds']['below'])
                ws.cell(row=row_idx, column=7, value=kr['thresholds']['meets'])
                ws.cell(row=row_idx, column=8, value=kr['thresholds']['good'])
                ws.cell(row=row_idx, column=9, value=kr['thresholds']['very_good'])
                ws.cell(row=row_idx, column=10, value=kr['thresholds']['exceptional'])
                ws.cell(row=row_idx, column=11, value=result['score'])
                ws.cell(row=row_idx, column=12, value=get_level_label(result['level']))

                # Apply color formatting to performance level cell
                level_cell = ws.cell(row=row_idx, column=12)
                level_cell.fill = PatternFill(start_color=colors[result['level']],
                                              end_color=colors[result['level']],
                                              fill_type='solid')
                level_cell.font = Font(bold=True, color='FFFFFF')
                level_cell.alignment = Alignment(horizontal='center', vertical='center')

                # Apply color formatting to score cell
                score_cell = ws.cell(row=row_idx, column=11)
                score_cell.fill = PatternFill(start_color=colors[result['level']],
                                              end_color=colors[result['level']],
                                              fill_type='solid')
                score_cell.font = Font(bold=True, color='FFFFFF')
                score_cell.alignment = Alignment(horizontal='center', vertical='center')

                row_idx += 1

            # Merge cells if there are multiple KRs
            if len(kr_list) > 1:
                end_row = row_idx - 1
                # Merge department name
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                # Merge objective name
                ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)

            # Apply formatting to department and objective cells
            dept_cell = ws.cell(row=start_row, column=1)
            dept_cell.alignment = Alignment(horizontal='center', vertical='center')
            dept_cell.font = Font(bold=True)

            obj_cell = ws.cell(row=start_row, column=2)
            obj_cell.alignment = Alignment(horizontal='center', vertical='center')
            obj_cell.font = Font(bold=True)

            # Add darker border after each objective (bottom of last row)
            end_row = row_idx - 1
            thick_bottom = Side(style='medium', color='000000')
            for col in range(1, 13):  # Columns A to L
                cell = ws.cell(row=end_row, column=col)
                # Preserve existing borders and add thick bottom
                cell.border = Border(
                    left=cell.border.left if cell.border else None,
                    right=cell.border.right if cell.border else None,
                    top=cell.border.top if cell.border else None,
                    bottom=thick_bottom
                )

    # Adjust column widths
    ws.column_dimensions['A'].width = 25  # Department
    ws.column_dimensions['B'].width = 30  # Objective
    ws.column_dimensions['C'].width = 30  # Key Result
    ws.column_dimensions['D'].width = 10  # Actual
    ws.column_dimensions['E'].width = 8   # Unit
    ws.column_dimensions['F'].width = 10  # Below
    ws.column_dimensions['G'].width = 10  # Meets
    ws.column_dimensions['H'].width = 15  # Good
    ws.column_dimensions['I'].width = 15  # Very Good
    ws.column_dimensions['J'].width = 10  # Exceptional
    ws.column_dimensions['K'].width = 10  # Score
    ws.column_dimensions['L'].width = 20  # Performance Level

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def main():
    st.set_page_config(page_title="OKR Tracker", page_icon="🎯", layout="wide")

    # Initialize
    if 'initialized' not in st.session_state:
        loaded_departments, loaded_lang = load_data()
        st.session_state.departments = loaded_departments
        st.session_state.language = loaded_lang
        st.session_state.new_krs = []
        st.session_state.initialized = True

    # Header
    col_title, col_view, col_lang = st.columns([3, 1, 1])
    with col_title:
        st.title(t("title"))
    with col_view:
        current_view = st.session_state.get('view_mode', 'full')
        col_grid, col_full = st.columns(2)
        with col_grid:
            if st.button(t("view_grid"), key="view_grid", type="primary" if current_view == 'grid' else "secondary", use_container_width=True):
                st.session_state.view_mode = 'grid'
                st.rerun()
        with col_full:
            if st.button(t("view_full"), key="view_full", type="primary" if current_view == 'full' else "secondary", use_container_width=True):
                st.session_state.view_mode = 'full'
                st.rerun()
    with col_lang:
        lang_options = {"en": "🇬🇧 English", "ru": "🇷🇺 Русский", "uz": "🇺🇿 Ўзбекча"}
        selected_lang = st.selectbox("", list(lang_options.keys()),
                                     format_func=lambda x: lang_options[x],
                                     index=list(lang_options.keys()).index(st.session_state.language),
                                     label_visibility="collapsed")
        if selected_lang != st.session_state.language:
            st.session_state.language = selected_lang
            save_data()
            st.rerun()

    # Legend
    st.markdown(f"### {t('performance_scale')}", unsafe_allow_html=True)
    cols = st.columns(5)
    for i, key in enumerate(["below", "meets", "good", "very_good", "exceptional"]):
        level = LEVELS[key]
        with cols[i]:
            pct_range = f"{score_to_percentage(level['min'])}%-{score_to_percentage(level['max'])}%"
            st.markdown(f'''
            <div style="background:{level['color']}; color:white; padding:12px; border-radius:8px; text-align:center;">
                <b>{get_level_label(key)}</b><br>
                <small>{level['min']:.2f}-{level['max']:.2f}</small><br>
                <small>({pct_range})</small>
            </div>
            ''', unsafe_allow_html=True)

    st.markdown("---")

    # Save/Load buttons
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        if st.button(t("save_data"), type="primary", use_container_width=True):
            save_data()
            st.success(t("data_saved"))
    with c2:
        if st.button(t("load_data"), use_container_width=True):
            dept, lang = load_data()
            if dept:
                st.session_state.departments = dept
                st.session_state.language = lang
                st.success(t("data_loaded"))
                st.rerun()
            else:
                st.warning(t("no_data"))
    with c3:
        if st.button(t("export_json"), use_container_width=True):
            export = []
            for dept in st.session_state.departments:
                dept_data = {"department": dept['name'], "objectives": []}
                for obj in dept.get('objectives', []):
                    obj_data = {"objective": obj['name'], "key_results": []}
                    scores = []
                    for kr in obj['key_results']:
                        res = calculate_score(kr['actual'], kr['metric_type'], kr['thresholds'])
                        scores.append(res['score'])
                        obj_data['key_results'].append({"name": kr['name'], "actual": kr['actual'],
                                                        "score": res['score'], "level": res['level']})
                    obj_data['average'] = round(sum(scores) / len(scores), 2) if scores else 0
                    obj_data['percentage'] = score_to_percentage(obj_data['average'])
                    dept_data['objectives'].append(obj_data)
                export.append(dept_data)
            st.json(export)
    with c4:
        excel_data = export_to_excel(st.session_state.departments)
        st.download_button(
            label=t("export_excel"),
            data=excel_data,
            file_name="okr_export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    st.markdown("---")

    # Create Department
    with st.expander(t("create_department"), expanded=len(st.session_state.departments) == 0):
        new_dept_name = st.text_input(t("department_name"), key="new_dept_name")
        if st.button(t("create_department") + " ✅", key="create_dept_btn"):
            if new_dept_name.strip():
                st.session_state.departments.append({
                    "id": str(uuid.uuid4()),
                    "name": new_dept_name.strip(),
                    "objectives": []
                })
                save_data()
                st.rerun()
            else:
                st.error("Please enter a department name")

    st.markdown("---")

    # Create Objective
    with st.expander(t("create_objective"), expanded=False):
        if not st.session_state.departments:
            st.warning(t("no_departments"))
        else:
            # Department selector
            dept_options = {dept['id']: dept['name'] for dept in st.session_state.departments}
            selected_dept_id = st.selectbox(
                t("select_department"),
                options=list(dept_options.keys()),
                format_func=lambda x: dept_options[x],
                key="selected_dept_for_obj"
            )

            new_obj_name = st.text_input(t("objective_name"), key="new_obj_name")

            st.markdown(f"#### {t('add_key_results')}")
            c1, c2, c3 = st.columns([3, 2, 1])
            with c1:
                kr_name = st.text_input(t("kr_name"), key="kr_name_input")
            with c2:
                kr_type = st.selectbox(t("type"), ["higher_better", "lower_better"],
                                       format_func=lambda x: t("higher_better") if x == "higher_better" else t(
                                           "lower_better"),
                                       key="kr_type_input")
            with c3:
                kr_unit = st.text_input(t("unit"), value="%", key="kr_unit_input")

            st.markdown(f"**{t('thresholds')}:**")
            t1, t2, t3, t4, t5 = st.columns(5)
            with t1:
                st.markdown(f"<small style='color:#d9534f;'>● 3.00</small>", unsafe_allow_html=True)
                th_below = st.number_input(t("below"), value=0.0, key="th_below")
            with t2:
                st.markdown(f"<small style='color:#f0ad4e;'>● 4.25</small>", unsafe_allow_html=True)
                th_meets = st.number_input(t("meets"), value=60.0, key="th_meets")
            with t3:
                st.markdown(f"<small style='color:#5cb85c;'>● 4.50</small>", unsafe_allow_html=True)
                th_good = st.number_input(t("good"), value=75.0, key="th_good")
            with t4:
                st.markdown(f"<small style='color:#28a745;'>● 4.75</small>", unsafe_allow_html=True)
                th_very_good = st.number_input(t("very_good"), value=90.0, key="th_very_good")
            with t5:
                st.markdown(f"<small style='color:#1e7b34;'>● 5.00</small>", unsafe_allow_html=True)
                th_exceptional = st.number_input(t("exceptional"), value=100.0, key="th_exceptional")

            if st.button(t("add_kr")):
                if kr_name.strip():
                    st.session_state.new_krs.append({
                        "id": str(uuid.uuid4()), "name": kr_name.strip(), "metric_type": kr_type,
                        "unit": kr_unit, "thresholds": {"below": th_below, "meets": th_meets,
                                                        "good": th_good, "very_good": th_very_good,
                                                        "exceptional": th_exceptional},
                        "actual": 0.0
                    })
                    st.rerun()

            if st.session_state.new_krs:
                st.markdown(f"**{t('added_krs')}:**")
                for i, kr in enumerate(st.session_state.new_krs):
                    col1, col2 = st.columns([5, 1])
                    with col1:
                        icon = "↑" if kr['metric_type'] == "higher_better" else "↓"
                        st.write(f"**KR{i + 1}: {kr['name']}** ({icon})")
                    with col2:
                        if st.button(f"❌", key=f"rm_{kr['id']}"):
                            st.session_state.new_krs = [k for k in st.session_state.new_krs if k['id'] != kr['id']]
                            st.rerun()

            if st.button(t("create"), type="primary"):
                if new_obj_name.strip() and st.session_state.new_krs:
                    # Find the department and add the objective
                    for dept in st.session_state.departments:
                        if dept['id'] == selected_dept_id:
                            dept['objectives'].append({
                                "id": str(uuid.uuid4()), "name": new_obj_name.strip(),
                                "key_results": st.session_state.new_krs.copy()
                            })
                            break
                    st.session_state.new_krs = []
                    save_data()
                    st.rerun()
                else:
                    st.error(t("enter_name_error"))

    st.markdown("---")

    # ===== DISPLAY OBJECTIVES =====
    if st.session_state.departments:
        view_mode = st.session_state.get('view_mode', 'full')

        # Iterate through departments
        for dept_idx, department in enumerate(st.session_state.departments):
            # Department header
            st.markdown(f"## 📁 {department['name']}")

            objectives = department.get('objectives', [])

            if not objectives:
                st.info(f"No objectives in this department yet.")
                # Add delete department button
                if st.button(t("delete_department"), key=f"del_dept_{department['id']}"):
                    st.session_state.departments = [d for d in st.session_state.departments if d['id'] != department['id']]
                    save_data()
                    st.rerun()
                st.markdown("---")
                continue

            if view_mode == 'grid':
                # GRID VIEW - 2 columns, compact cards
                for row_start in range(0, len(objectives), 2):
                    cols = st.columns(2)
                    for col_idx in range(2):
                        obj_idx = row_start + col_idx
                        if obj_idx < len(objectives):
                            objective = objectives[obj_idx]
                            krs = objective.get('key_results', [])

                            if not krs:
                                with cols[col_idx]:
                                    st.warning(t("no_krs"))
                                continue

                            # Calculate scores
                            results = [calculate_score(kr['actual'], kr['metric_type'], kr['thresholds']) for kr in krs]
                            avg_score = sum(r['score'] for r in results) / len(results)
                            avg_level = get_level_for_score(avg_score)
                            avg_pct = score_to_percentage(avg_score)

                            with cols[col_idx]:
                                # Compact header
                                st.markdown(f'''
                                <div style="background:{avg_level['color']}; color:white; padding:8px 12px;
                                            border-radius:6px; margin-bottom:8px; font-weight:bold; font-size:13px;">
                                    📋 {objective['name']} | Avg: {avg_score:.2f}
                                </div>
                                ''', unsafe_allow_html=True)

                                # Compact gauge
                                gauge_html = create_gauge(avg_score, compact=True)
                                components.html(gauge_html, height=170)

                                # Compact KR table
                                kr_rows = "".join([
                                    f'<tr><td style="padding:3px 6px;font-size:11px;">KR{i + 1}</td>'
                                    f'<td style="padding:3px 6px;font-size:11px;">{kr["name"][:18]}{"..." if len(kr["name"]) > 18 else ""}</td>'
                                    f'<td style="padding:3px 6px;font-size:11px;text-align:center;">{kr["actual"]}{kr["unit"]}</td>'
                                    f'<td style="padding:3px 6px;font-size:11px;text-align:center;background:{results[i]["level_info"]["color"]};color:white;">{results[i]["score"]:.2f}</td></tr>'
                                    for i, kr in enumerate(krs)
                                ])
                                components.html(f'''
                                <table style="width:100%;border-collapse:collapse;font-size:11px;">
                                    <tr style="background:#4472C4;color:white;">
                                        <th style="padding:4px;">KR</th><th style="padding:4px;">Name</th>
                                        <th style="padding:4px;">Fact</th><th style="padding:4px;">Score</th>
                                    </tr>
                                    {kr_rows}
                                </table>
                                ''', height=30 + len(krs) * 24)

                                # Edit expander
                                with st.expander("✏️ Edit"):
                                    for kr_idx, kr in enumerate(krs):
                                        new_val = st.number_input(f"KR{kr_idx + 1}", value=float(kr['actual']),
                                                                  key=f"grid_d{dept_idx}_o{obj_idx}_{kr['id']}")
                                        if new_val != kr['actual']:
                                            st.session_state.departments[dept_idx]['objectives'][obj_idx]['key_results'][kr_idx]['actual'] = new_val
                                            save_data()
                                            st.rerun()
                                    if st.button("🗑️ Delete", key=f"del_grid_d{dept_idx}_{objective['id']}"):
                                        st.session_state.departments[dept_idx]['objectives'] = [o for o in objectives if
                                                                       o['id'] != objective['id']]
                                        save_data()
                                        st.rerun()
            else:
                # FULL VIEW - Original layout
                for obj_idx, objective in enumerate(objectives):
                    krs = objective.get('key_results', [])

                    if not krs:
                        st.warning(t("no_krs"))
                        continue

                    # Calculate scores
                    results = []
                    for kr in krs:
                        result = calculate_score(kr['actual'], kr['metric_type'], kr['thresholds'])
                        results.append(result)

                    avg_score = sum(r['score'] for r in results) / len(results)
                    avg_level = get_level_for_score(avg_score)
                    avg_pct = score_to_percentage(avg_score)

                    # Objective title
                    st.markdown(f"### {objective['name']}")

                    # Yellow header bar
                    st.markdown(f'''
                    <div style="background:#FFC000; padding:12px 20px; border-radius:8px;
                                display:flex; justify-content:space-between; align-items:center; margin-bottom:15px;">
                        <span style="font-weight:bold; font-size:16px;">📋 {objective['name']}</span>
                        <span style="background:{avg_level['color']}; color:white; padding:6px 18px;
                                     border-radius:20px; font-weight:bold; font-size:16px;">
                            Avg: {avg_score:.2f}
                        </span>
                    </div>
                    ''', unsafe_allow_html=True)

                    st.markdown(f"**Results**", unsafe_allow_html=True)
                    with st.expander(f"{objective['name']}", expanded=True):
                        col_table, col_gauge = st.columns([3, 1])

                        with col_table:
                            # Build DataFrame for editable table (KR, Key Result, Fact, Score only)
                            table_data = []
                            for kr_idx, kr in enumerate(krs):
                                result = results[kr_idx]
                                table_data.append({
                                    "KR": f"KR{kr_idx + 1}",
                                    t("key_result"): kr['name'],
                                    t("fact"): kr['actual'],
                                    "Score": result['score'],
                                })

                            df = pd.DataFrame(table_data)

                            # Editable table for Fact column
                            edited_df = st.data_editor(
                                df,
                                column_config={
                                    "KR": st.column_config.TextColumn("KR", disabled=True, width="small"),
                                    t("key_result"): st.column_config.TextColumn(t("key_result"), disabled=True,
                                                                                 width="medium"),
                                    t("fact"): st.column_config.NumberColumn(t("fact"), min_value=-1000, max_value=10000,
                                                                             step=1, format="%.1f"),
                                    "Score": st.column_config.NumberColumn("Score", disabled=True, format="%.2f",
                                                                           width="small"),
                                },
                                hide_index=True,
                                use_container_width=True,
                                key=f"editor_d{dept_idx}_o{obj_idx}_{objective['id']}"
                            )

                            # Update actual values from edited dataframe
                            for i, row in edited_df.iterrows():
                                if i < len(krs):
                                    new_actual = row[t("fact")]
                                    if new_actual != krs[i]['actual']:
                                        st.session_state.departments[dept_idx]['objectives'][obj_idx]['key_results'][i]['actual'] = new_actual
                                        save_data()
                                        st.rerun()

                            # ===== VISUAL HTML TABLE (Results Breakdown) =====
                            st.markdown(f"#### {t('result')}s Breakdown")

                            html_table = f'''
                            <table style="width:100%; border-collapse:collapse; font-size:13px; margin-top:10px;">
                                <thead>
                                    <tr style="background:#4472C4; color:white;">
                                        <th style="padding:10px; border:1px solid #2F5496;">KR</th>
                                        <th style="padding:10px; border:1px solid #2F5496;">{t('key_result')}</th>
                                        <th style="padding:10px; border:1px solid #2F5496;">{t('fact')}</th>
                                        <th style="padding:10px; border:1px solid #2F5496; background:#d9534f;">{get_level_label('below')}<br><small>3.00</small></th>
                                        <th style="padding:10px; border:1px solid #2F5496; background:#f0ad4e; color:#000;">{get_level_label('meets')}<br><small>4.25</small></th>
                                        <th style="padding:10px; border:1px solid #2F5496; background:#5cb85c;">{get_level_label('good')}<br><small>4.50</small></th>
                                        <th style="padding:10px; border:1px solid #2F5496; background:#28a745;">{get_level_label('very_good')}<br><small>4.75</small></th>
                                        <th style="padding:10px; border:1px solid #2F5496; background:#1e7b34;">{get_level_label('exceptional')}<br><small>5.00</small></th>
                                        <th style="padding:10px; border:1px solid #2F5496;">{t('result')}</th>
                                        <th style="padding:10px; border:1px solid #2F5496;">{t('delete')}</th>
                                    </tr>
                                </thead>
                                <tbody>
                            '''

                            for kr_idx, kr in enumerate(krs):
                                result = results[kr_idx]
                                th = kr['thresholds']
                                level = result['level']

                                # Cell highlighting
                                cells = {
                                    'below': '' if level != 'below' else 'background:#d9534f; color:white; font-weight:bold;',
                                    'meets': '' if level != 'meets' else 'background:#f0ad4e; color:#000; font-weight:bold;',
                                    'good': '' if level != 'good' else 'background:#5cb85c; color:white; font-weight:bold;',
                                    'very_good': '' if level != 'very_good' else 'background:#28a745; color:white; font-weight:bold;',
                                    'exceptional': '' if level != 'exceptional' else 'background:#1e7b34; color:white; font-weight:bold;',
                                }

                                if kr['metric_type'] == "higher_better":
                                    th_texts = [f"<{th['below']}", f"≥{th['meets']}", f"≥{th['good']}",
                                                f"≥{th['very_good']}", f"≥{th['exceptional']}"]
                                else:
                                    th_texts = [f">{th['below']}", f"≤{th['meets']}", f"≤{th['good']}",
                                                f"≤{th['very_good']}", f"≤{th['exceptional']}"]

                                row_bg = '#F8F9FA' if kr_idx % 2 == 0 else '#FFFFFF'

                                html_table += f'''
                                    <tr style="background:{row_bg};">
                                        <td style="padding:10px; border:1px solid #ddd; font-weight:bold;">KR{kr_idx + 1}</td>
                                        <td style="padding:10px; border:1px solid #ddd; text-align:left;">{kr['name']}</td>
                                        <td style="padding:10px; border:1px solid #ddd; background:#E2EFDA; font-weight:bold;">{kr['actual']}{kr['unit']}</td>
                                        <td style="padding:10px; border:1px solid #ddd; {cells['below']}">{th_texts[0]}</td>
                                        <td style="padding:10px; border:1px solid #ddd; {cells['meets']}">{th_texts[1]}</td>
                                        <td style="padding:10px; border:1px solid #ddd; {cells['good']}">{th_texts[2]}</td>
                                        <td style="padding:10px; border:1px solid #ddd; {cells['very_good']}">{th_texts[3]}</td>
                                        <td style="padding:10px; border:1px solid #ddd; {cells['exceptional']}">{th_texts[4]}</td>
                                        <td style="padding:10px; border:1px solid #ddd; background:{result['level_info']['color']}; color:white; font-weight:bold;">{result['score']:.2f}</td>
                                        <td style="padding:10px; border:1px solid #ddd;">🗑️</td>
                                    </tr>
                                '''

                            # Formula row
                            kr_formula = " + ".join([f"KR{i + 1}" for i in range(len(krs))])
                            html_table += f'''
                                    <tr style="background:#FFF2CC; font-weight:bold;">
                                        <td colspan="8" style="padding:12px; border:2px solid #BF9000; text-align:right;">
                                            ({kr_formula}) / {len(krs)} =
                                        </td>
                                        <td colspan="2" style="padding:12px; border:2px solid #BF9000; background:{avg_level['color']}; color:white; font-size:16px;">
                                            {avg_score:.2f}
                                        </td>
                                    </tr>
                                </tbody>
                            </table>
                            '''

                            # Render HTML table
                            table_height = 80 + (len(krs) * 50) + 60
                            components.html(html_table, height=table_height, scrolling=False)

                            # Delete KR buttons
                            st.markdown(f"#### {t('delete_krs')}")
                            del_cols = st.columns(len(krs) + 1)
                            for kr_idx, kr in enumerate(krs):
                                with del_cols[kr_idx]:
                                    if st.button(f"{t('delete')} KR{kr_idx + 1}",
                                                 key=f"del_kr_d{dept_idx}_o{obj_idx}_{kr['id']}"):
                                        st.session_state.departments[dept_idx]['objectives'][obj_idx]['key_results'] = [
                                            k for k in krs if k['id'] != kr['id']]
                                        save_data()
                                        st.rerun()

                        with col_gauge:
                            st.markdown(f"### {t('score')}")
                            gauge_html = create_gauge(avg_score)
                            components.html(gauge_html, height=300)

                            # Level indicator
                            st.markdown(f'''
                            <div style="text-align:center; margin-top:10px;">
                                <div style="background:{avg_level['color']}; color:white; padding:15px;
                                            border-radius:10px; font-size:20px; font-weight:bold;">
                                    {get_level_label(avg_level['key'])}<br>
                                    <small>{avg_score:.2f} ({avg_pct}%)</small>
                                </div>
                            </div>
                            ''', unsafe_allow_html=True)

                        # Add KR to this objective
                        with st.expander(t("add_kr_to_obj")):
                            ac1, ac2, ac3 = st.columns([3, 2, 1])
                            with ac1:
                                add_name = st.text_input(t("kr_name"), key=f"add_name_d{dept_idx}_o{obj_idx}")
                            with ac2:
                                add_type = st.selectbox(t("type"), ["higher_better", "lower_better"],
                                                        format_func=lambda x: "↑" if x == "higher_better" else "↓",
                                                        key=f"add_type_d{dept_idx}_o{obj_idx}")
                            with ac3:
                                add_unit = st.text_input(t("unit"), value="%", key=f"add_unit_d{dept_idx}_o{obj_idx}")

                            at1, at2, at3, at4, at5 = st.columns(5)
                            with at1:
                                st.markdown(f"<small style='color:#d9534f;'>● 3.00</small>", unsafe_allow_html=True)
                                add_below = st.number_input(t("below"), value=0.0, key=f"add_below_d{dept_idx}_o{obj_idx}")
                            with at2:
                                st.markdown(f"<small style='color:#f0ad4e;'>● 4.25</small>", unsafe_allow_html=True)
                                add_meets = st.number_input(t("meets"), value=60.0, key=f"add_meets_d{dept_idx}_o{obj_idx}")
                            with at3:
                                st.markdown(f"<small style='color:#5cb85c;'>● 4.50</small>", unsafe_allow_html=True)
                                add_good = st.number_input(t("good"), value=75.0,
                                                            key=f"add_good_d{dept_idx}_o{obj_idx}")
                            with at4:
                                st.markdown(f"<small style='color:#28a745;'>● 4.75</small>", unsafe_allow_html=True)
                                add_very_good = st.number_input(t("very_good"), value=90.0, key=f"add_very_good_d{dept_idx}_o{obj_idx}")
                            with at5:
                                st.markdown(f"<small style='color:#1e7b34;'>● 5.00</small>", unsafe_allow_html=True)
                                add_exc = st.number_input(t("exceptional"), value=100.0, key=f"add_exc_d{dept_idx}_o{obj_idx}")

                            if st.button(t("add"), key=f"add_btn_d{dept_idx}_o{obj_idx}"):
                                if add_name.strip():
                                    st.session_state.departments[dept_idx]['objectives'][obj_idx]['key_results'].append({
                                        "id": str(uuid.uuid4()), "name": add_name.strip(), "metric_type": add_type,
                                        "unit": add_unit, "thresholds": {"below": add_below, "meets": add_meets,
                                                                         "good": add_good, "very_good": add_very_good,
                                                                         "exceptional": add_exc},
                                        "actual": 0.0
                                    })
                                    save_data()
                                    st.rerun()

                        # Delete objective
                        if st.button(f"{t('delete_objective')} '{objective['name']}'", key=f"del_obj_d{dept_idx}_{objective['id']}"):
                            st.session_state.departments[dept_idx]['objectives'] = [o for o in objectives if
                                                           o['id'] != objective['id']]
                            save_data()
                            st.rerun()

                    st.markdown("---")

            # Add delete department button at end of each department
            if st.button(t("delete_department") + f" '{department['name']}'", key=f"del_dept_end_{department['id']}"):
                st.session_state.departments = [d for d in st.session_state.departments if d['id'] != department['id']]
                save_data()
                st.rerun()

            st.markdown("---")

    else:
        st.info(t("create_first"))
        if st.button(t("load_demo")):
            # Create departments with demo objectives
            st.session_state.departments = [{
                "id": str(uuid.uuid4()),
                "name": "PMO - Project Management Office",
                "objectives": [
                # Цель 1: Обеспечить своевременную реализацию проектов (20%)
                {
                    "id": str(uuid.uuid4()),
                    "name": "Цель 1: Обеспечить своевременную реализацию проектов",
                    "weight": 20,
                    "key_results": [
                        {"id": str(uuid.uuid4()), "name": "KR1.1 Проекты завершенные в срок (% от кол-ва проектов)",
                         "metric_type": "higher_better", "unit": "%",
                         "thresholds": {"below": 50, "meets": 60, "good": 80, "very_good": 100, "exceptional": 120}, "actual": 0},
                        {"id": str(uuid.uuid4()), "name": "KR1.2 Задачи в JIRA, завершенные в срок (%)",
                         "metric_type": "higher_better", "unit": "%",
                         "thresholds": {"below": 50, "meets": 65, "good": 95, "very_good": 100, "exceptional": 200}, "actual": 0},
                        {"id": str(uuid.uuid4()), "name": "KR1.3 Переносы сроков заверш задач в JIRA (% от общего кол-ва)",
                         "metric_type": "lower_better", "unit": "%",
                         "thresholds": {"below": 30, "meets": 20, "good": 15, "very_good": 5, "exceptional": 0}, "actual": 0},
                    ]
                },
                # Цель 2: Управление рисками и бюджетом проектов (20%)
                {
                    "id": str(uuid.uuid4()),
                    "name": "Цель 2: Управление рисками и бюджетом проектов",
                    "weight": 20,
                    "key_results": [
                        {"id": str(uuid.uuid4()), "name": "KR2.1 Проекты в рамках бюджетов (% без превышения)",
                         "metric_type": "higher_better", "unit": "%",
                         "thresholds": {"below": 50, "meets": 60, "good": 75, "very_good": 90, "exceptional": 100}, "actual": 0},
                        {"id": str(uuid.uuid4()), "name": "KR2.2 Неучтенные риски возникшие после начала проекта (кол-во)",
                         "metric_type": "lower_better", "unit": "",
                         "thresholds": {"below": 10, "meets": 5, "good": 2, "very_good": 1, "exceptional": 0}, "actual": 0},
                        {"id": str(uuid.uuid4()), "name": "KR2.3 Повысить точность оценки трудозатрат до 75%",
                         "metric_type": "higher_better", "unit": "%",
                         "thresholds": {"below": 50, "meets": 75, "good": 80, "very_good": 85, "exceptional": 100}, "actual": 0},
                        {"id": str(uuid.uuid4()), "name": "KR2.4 Процент рисков с планами митигации (%)",
                         "metric_type": "higher_better", "unit": "%",
                         "thresholds": {"below": 20, "meets": 50, "good": 60, "very_good": 80, "exceptional": 100}, "actual": 0},
                    ]
                },
                # Цель 3: Управление качеством и отчетность (20%)
                {
                    "id": str(uuid.uuid4()),
                    "name": "Цель 3: Управление качеством и отчетность",
                    "weight": 20,
                    "key_results": [
                        {"id": str(uuid.uuid4()), "name": "KR3.1 Своевременность отчетов W,Q,Y, другие (задержка, дней)",
                         "metric_type": "lower_better", "unit": " дней",
                         "thresholds": {"below": 5, "meets": 3, "good": 2, "very_good": 1, "exceptional": 0}, "actual": 0},
                        {"id": str(uuid.uuid4()), "name": "KR3.2 Уровень использования ресурсов (resource utilization) %",
                         "metric_type": "higher_better", "unit": "%",
                         "thresholds": {"below": 75, "meets": 85, "good": 90, "very_good": 95, "exceptional": 100}, "actual": 0},
                        {"id": str(uuid.uuid4()), "name": "KR3.3 Реагирование на изменения (Response time to changes) часы",
                         "metric_type": "lower_better", "unit": " часов",
                         "thresholds": {"below": 5, "meets": 3, "good": 2, "very_good": 1, "exceptional": 0}, "actual": 0},
                        {"id": str(uuid.uuid4()), "name": "KR3.4 Среднее время от инициации до завершения проекта (нед)",
                         "metric_type": "lower_better", "unit": " нед",
                         "thresholds": {"below": 10, "meets": 8, "good": 6, "very_good": 5, "exceptional": 4}, "actual": 0},
                    ]
                },
                # Цель 4: Усиление состава и человеческий капитал (10%)
                {
                    "id": str(uuid.uuid4()),
                    "name": "Цель 4: Усиление состава и человеческий капитал",
                    "weight": 10,
                    "key_results": [
                        {"id": str(uuid.uuid4()), "name": "KR4.1 Комплектация штата (6 свободных вакансий в штате)",
                         "metric_type": "higher_better", "unit": "",
                         "thresholds": {"below": 2, "meets": 3, "good": 4, "very_good": 5, "exceptional": 6}, "actual": 0},
                        {"id": str(uuid.uuid4()), "name": "KR4.2 Набор и подготовка стажеров (16 вакансий)",
                         "metric_type": "higher_better", "unit": "",
                         "thresholds": {"below": 3, "meets": 6, "good": 10, "very_good": 12, "exceptional": 16}, "actual": 0},
                        {"id": str(uuid.uuid4()), "name": "KR4.3 % сотрудников с ростом оклада, должности",
                         "metric_type": "higher_better", "unit": "%",
                         "thresholds": {"below": 0, "meets": 10, "good": 20, "very_good": 30, "exceptional": 40}, "actual": 0},
                    ]
                },
                # Цель 5: Улучшение продуктов (10%)
                {
                    "id": str(uuid.uuid4()),
                    "name": "Цель 5: Улучшение продуктов",
                    "weight": 10,
                    "key_results": [
                        {"id": str(uuid.uuid4()), "name": "KR5.1 Увеличить долю проектов, связанных со стратегическими целями Банка, до 85%",
                         "metric_type": "higher_better", "unit": "%",
                         "thresholds": {"below": 75, "meets": 85, "good": 90, "very_good": 95, "exceptional": 100}, "actual": 0},
                        {"id": str(uuid.uuid4()), "name": "KR5.2 % продуктов с повторными багами (Defect/error rate)",
                         "metric_type": "lower_better", "unit": "%",
                         "thresholds": {"below": 20, "meets": 15, "good": 10, "very_good": 5, "exceptional": 0}, "actual": 0},
                        {"id": str(uuid.uuid4()), "name": "KR5.3 Обеспечить участие 100% членов команды в обучении по Agile/Scrum",
                         "metric_type": "higher_better", "unit": "%",
                         "thresholds": {"below": 80, "meets": 90, "good": 95, "very_good": 100, "exceptional": 100}, "actual": 0},
                        {"id": str(uuid.uuid4()), "name": "KR5.4 Провести 6 внутренних воркшопов по методологиям и новым технологиям",
                         "metric_type": "higher_better", "unit": "",
                         "thresholds": {"below": 4, "meets": 6, "good": 7, "very_good": 8, "exceptional": 9}, "actual": 0},
                    ]
                },
                # Цель 6: Системная и бизнес аналитика и ее автоматизация (20%)
                {
                    "id": str(uuid.uuid4()),
                    "name": "Цель 6: Системная и бизнес аналитика и ее автоматизация",
                    "weight": 20,
                    "key_results": [
                        {"id": str(uuid.uuid4()), "name": "KR6.1 Уровень автоматизации процессов проектного управления",
                         "metric_type": "higher_better", "unit": "%",
                         "thresholds": {"below": 75, "meets": 85, "good": 90, "very_good": 95, "exceptional": 100}, "actual": 0},
                        {"id": str(uuid.uuid4()), "name": "KR6.2 Качество описание бизнес процессов (изменение BPMN) %",
                         "metric_type": "lower_better", "unit": "%",
                         "thresholds": {"below": 20, "meets": 15, "good": 10, "very_good": 5, "exceptional": 0}, "actual": 0},
                        {"id": str(uuid.uuid4()), "name": "KR6.3 Процент изменений плана проекта после планирования",
                         "metric_type": "lower_better", "unit": "%",
                         "thresholds": {"below": 20, "meets": 15, "good": 10, "very_good": 5, "exceptional": 0}, "actual": 0},
                    ]
                },
                ]
            }]
            save_data()
            st.rerun()


if __name__ == "__main__":
    main()